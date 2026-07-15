from werkzeug.utils import secure_filename  # 用于安全处理上传文件名
from bjc import sf_db, dui_db  # 导入数据库相关操作
from datetime import datetime  # 处理日期和时间
import pandas as pd  # 处理数据库中导出的相关数据
from message_service import MessageService  # 导入统一消息发送服务
from config import *  # 从配置文件中导入所有应用
from department_permissions import permission_manager
import os
from io import BytesIO  # 用于操作文件和目录
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_from_directory, \
    send_file  # 导入Flask核心模块，用于web开发，包括模块渲染，请求处理，json响应
import re
import json
import urllib.request
import subprocess
import sys
import threading
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import quote

app = Flask(__name__)  # 创建 Flask 应用实例，__name__ 用于定位模板和静态文件
app.secret_key = APP_CONFIG['secret_key']  # 设置秘钥，暂时没用到
app.config['UPLOAD_FOLDER'] = APP_CONFIG['upload_folder']  # 设置文件保存路径，图片的路径
app.config['MAX_CONTENT_LENGTH'] = APP_CONFIG[
    'max_content_length']  # 配置文件大小限制，最大16m                                      #配置文件大小限制，最大16m
LIKE_IMAGE_FOLDER = r"D:\tuchuangai\点赞图片"
INNOVATION_STAR_MEDIA_FOLDER = os.path.abspath(
    os.environ.get('INNOVATION_STAR_MEDIA_FOLDER', r'D:\tuchuangai\创新星主场')
)
INNOVATION_STAR_EDITOR_NAMES = {'韩雅俊'}
INNOVATION_STAR_IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.webp'}
INNOVATION_STAR_VIDEO_EXTENSIONS = {'.mp4', '.webm', '.ogg', '.mov', '.m4v'}
OPERATION_DEPARTMENTS = {'运营一部', '运营二部', '运营三部', '运营六部', '运营七部'}
OPERATION_WATCHER_NAMES = ['孙洁', '侯梁']


def _safe_print(*args, **kwargs):
    try:
        print(*args, **kwargs)
    except Exception:
        pass


def _database_upload_path(filename):
    """Keep database paths independent from the physical storage drive."""
    return f"uploads/{filename}".replace('\\', '/')


def _innovation_debug_report(hypothesis_id, location, msg, data=None):
    try:
        debug_url = 'http://127.0.0.1:7777/event'
        debug_session_id = 'innovation-comment-notify'
        env_path = os.path.join('.dbg', 'innovation-comment-notify.env')
        try:
            with open(env_path, 'r', encoding='utf-8') as env_file:
                env_content = env_file.read()
            for line in env_content.splitlines():
                if line.startswith('DEBUG_SERVER_URL='):
                    debug_url = line.split('=', 1)[1].strip() or debug_url
                elif line.startswith('DEBUG_SESSION_ID='):
                    debug_session_id = line.split('=', 1)[1].strip() or debug_session_id
        except Exception:
            pass
        payload = {
            'sessionId': debug_session_id,
            'runId': 'post-fix',
            'hypothesisId': hypothesis_id,
            'location': location,
            'msg': f'[DEBUG] {msg}',
            'data': data or {}
        }
        urllib.request.urlopen(
            urllib.request.Request(
                debug_url,
                data=json.dumps(payload, ensure_ascii=False).encode('utf-8'),
                headers={'Content-Type': 'application/json'}
            ),
            timeout=2
        ).read()
    except Exception:
        pass


for folder in [APP_CONFIG['upload_folder'], APP_CONFIG['export_folder']]:  # 确保必要目录存在
    os.makedirs(folder, exist_ok=True)
try:
    os.makedirs(LIKE_IMAGE_FOLDER, exist_ok=True)
except Exception:
    LIKE_IMAGE_FOLDER = os.path.join(APP_CONFIG['upload_folder'], 'like_images')
    os.makedirs(LIKE_IMAGE_FOLDER, exist_ok=True)


def get_message_service():  # 获取消息发送服务的函数
    """根据请求参数获取对应公司的消息服务实例"""
    company_key = request.args.get('company', 'company1')  # 默认使用company1
    return MessageService(company_key)  # 返回对应公司消息服务实例


ADMIN_USERS = ['admin', '管理员', '系统管理员']  # 管理员用户列表，暂时没用到


def is_admin(username):  # 检查是否管理员，暂时无法查询
    """检查用户是否为管理员"""
    return username in ADMIN_USERS  # 若在返回TRUE，反之FALSE


def _get_current_feishu_user_name():
    raw_name = session.get('feishu_user_name', '')
    name_parts = str(raw_name).split('（', 1) if raw_name else []
    return name_parts[0].strip() if name_parts else str(raw_name).strip()


def _innovation_lookup_feishu_id_by_name(user_name):
    name = str(user_name or '').strip()
    if not name:
        return ''
    try:
        rows = sf_db(
            f"""
            SELECT TOP 1 FeiShu_ID
            FROM feishu_id
            WHERE YONGHU = N'{name.replace("'", "''")}'
            ORDER BY CASE WHEN FeiShu_ID LIKE 'ou[_]%%' THEN 0 ELSE 1 END, FeiShu_ID
            """
        ) or []
    except Exception:
        rows = []
    if not rows:
        return ''
    first = rows[0]
    if isinstance(first, dict):
        return str(first.get('FeiShu_ID') or first.get('feishu_id') or '').strip()
    if isinstance(first, (list, tuple)):
        return str(first[0] if first else '').strip()
    return str(first or '').strip()


def _innovation_get_user_departments_by_name(user_name):
    feishu_id = _innovation_lookup_feishu_id_by_name(user_name)
    if not feishu_id:
        return []
    try:
        rows = permission_manager.get_user_departments(feishu_id) or []
    except Exception:
        rows = []
    out = []
    seen = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        if str(row.get('status') or '').strip() in {'invalid', 'unmapped'}:
            continue
        dept_name = str(row.get('name') or '').strip()
        if not dept_name or dept_name in seen:
            continue
        seen.add(dept_name)
        out.append(dept_name)
    return out


def _innovation_get_project_basic_info(project_id):
    try:
        rows = sf_db(
            f"""
            SELECT TOP 1
                ISNULL(CAST([标题] AS NVARCHAR(500)), ''),
                ISNULL(CAST([发起人] AS NVARCHAR(200)), '')
            FROM chuangxin_tibao1
            WHERE [编号] = {int(project_id)}
            """
        ) or []
    except Exception as e:
        _safe_print(f"⚠️ 查询创新项目基础信息失败: id={project_id}, error={e}")
        return {'title': '', 'initiator': ''}
    if not rows:
        return {'title': '', 'initiator': ''}
    first = rows[0]
    if isinstance(first, dict):
        return {
            'title': str(first.get('标题') or first.get('title') or '').strip(),
            'initiator': str(first.get('发起人') or first.get('initiator') or '').strip(),
        }
    if isinstance(first, (list, tuple)):
        return {
            'title': str(first[0] if len(first) > 0 and first[0] is not None else '').strip(),
            'initiator': str(first[1] if len(first) > 1 and first[1] is not None else '').strip(),
        }
    return {'title': '', 'initiator': ''}


def _innovation_notify_initiator_on_comment(project_id, commenter_name, comment_text='', has_image=False):
    # #region debug-point B:notify-entry
    _innovation_debug_report('B', 'innovation.web_app:_innovation_notify_initiator_on_comment:entry', 'enter notify helper', {'project_id': project_id, 'commenter': str(commenter_name or '').strip(), 'has_comment': bool(str(comment_text or '').strip()), 'has_image': bool(has_image)})
    # #endregion
    commenter = str(commenter_name or '').strip()
    if not project_id or not commenter:
        # #region debug-point A:missing-args
        _innovation_debug_report('A', 'innovation.web_app:_innovation_notify_initiator_on_comment:missing_args', 'skip notify because args missing', {'project_id': project_id, 'commenter': commenter})
        # #endregion
        return {'sent': False, 'reason': 'missing_args'}
    project_info = _innovation_get_project_basic_info(project_id)
    initiator = str(project_info.get('initiator') or '').strip()
    title = str(project_info.get('title') or '').strip()
    # #region debug-point B:project-info
    _innovation_debug_report('B', 'innovation.web_app:_innovation_notify_initiator_on_comment:project_info', 'loaded project info for notify', {'project_id': project_id, 'initiator': initiator, 'title': title})
    # #endregion
    if not initiator:
        # #region debug-point B:initiator-missing
        _innovation_debug_report('B', 'innovation.web_app:_innovation_notify_initiator_on_comment:initiator_missing', 'skip notify because initiator not found', {'project_id': project_id, 'commenter': commenter})
        # #endregion
        return {'sent': False, 'reason': 'initiator_not_found'}
    if initiator == commenter:
        # #region debug-point E:self-comment
        _innovation_debug_report('E', 'innovation.web_app:_innovation_notify_initiator_on_comment:self_comment', 'initiator equals commenter but notify will continue', {'project_id': project_id, 'initiator': initiator, 'commenter': commenter})
        # #endregion

    comment_preview = str(comment_text or '').strip()
    if comment_preview:
        comment_preview = re.sub(r'\s+', ' ', comment_preview)
        if len(comment_preview) > 120:
            comment_preview = comment_preview[:120] + '...'
    elif has_image:
        comment_preview = '评论中附带了图片'
    else:
        comment_preview = '你的提案收到了新的评论'

    manage_url = "http://223.78.73.100:8000/innovation/manage"
    message = (
        "💬 你的创新提案收到了新评论\n\n"
        f"提案编号：{project_id}\n"
        f"提案标题：{title or '未命名提案'}\n"
        f"评论人：{commenter}\n"
        f"评论内容：{comment_preview}\n\n"
        f"请前往管理页查看：\n{manage_url}"
    )
    try:
        # #region debug-point C:send-attempt
        _innovation_debug_report('C', 'innovation.web_app:_innovation_notify_initiator_on_comment:send_attempt', 'attempt send comment notify', {'project_id': project_id, 'initiator': initiator, 'commenter': commenter})
        # #endregion
        ok = bool(get_message_service().send_message(initiator, message))
        # #region debug-point C:send-result
        _innovation_debug_report('C', 'innovation.web_app:_innovation_notify_initiator_on_comment:send_result', 'comment notify send finished', {'project_id': project_id, 'initiator': initiator, 'commenter': commenter, 'ok': ok})
        # #endregion
        if not ok:
            _safe_print(f"⚠️ 创新评论通知发送失败: id={project_id}, initiator={initiator}, commenter={commenter}")
        return {'sent': ok, 'reason': '' if ok else 'send_failed', 'initiator': initiator}
    except Exception as e:
        # #region debug-point C:send-error
        _innovation_debug_report('C', 'innovation.web_app:_innovation_notify_initiator_on_comment:send_error', 'comment notify send raised exception', {'project_id': project_id, 'initiator': initiator, 'commenter': commenter, 'error': str(e)})
        # #endregion
        _safe_print(f"⚠️ 创新评论通知发送异常: id={project_id}, initiator={initiator}, commenter={commenter}, error={e}")
        return {'sent': False, 'reason': str(e), 'initiator': initiator}


def _innovation_collect_operation_watcher_ids(initiator_name, dept_list):
    watcher_ids = []
    watcher_seen = set()
    initiator_departments = _innovation_get_user_departments_by_name(initiator_name)
    related_departments = set([str(x or '').strip() for x in list(dept_list or []) + list(initiator_departments or []) if str(x or '').strip()])
    matched_operation_departments = sorted(list(related_departments & OPERATION_DEPARTMENTS))
    if not matched_operation_departments:
        return [], initiator_departments, matched_operation_departments
    for watcher_name in OPERATION_WATCHER_NAMES:
        watcher_id = _innovation_lookup_feishu_id_by_name(watcher_name)
        if not watcher_id or watcher_id in watcher_seen:
            continue
        watcher_seen.add(watcher_id)
        watcher_ids.append(watcher_id)
    return watcher_ids, initiator_departments, matched_operation_departments


def _escape_sql_literal_for_pytds(value):
    """转义 SQL 字面量，兼容 pytds 对 % 的格式化处理。"""
    if value is None:
        return ''
    return str(value).replace("'", "''").replace("%", "%%")


def _ensure_nvarchar_max_column(table_name, column_name):
    """确保指定文本列可容纳长文本，避免提交长内容时被截断。"""
    try:
        sql = f"""
            SELECT DATA_TYPE, CHARACTER_MAXIMUM_LENGTH
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME='{table_name}' AND COLUMN_NAME=N'{column_name}'
        """
        rows = sf_db(sql) or []
        if not rows:
            return False
        data_type = str(rows[0][0] or '').strip().lower()
        max_length = rows[0][1]
        if data_type == 'nvarchar' and int(max_length or -1) == -1:
            return True
        if data_type == 'ntext':
            return True
        dui_db(f"ALTER TABLE {table_name} ALTER COLUMN [{column_name}] NVARCHAR(MAX) NULL")
        return True
    except Exception as e:
        _safe_print(f"⚠️ 扩展字段容量失败: {table_name}.{column_name} -> {e}")
        return False


def _is_mobile_request():
    ua = (request.headers.get('User-Agent') or '').lower()
    mobile_keywords = ['iphone', 'ipad', 'android', 'mobile', 'harmony', 'micromessenger']
    return any(k in ua for k in mobile_keywords)


def _innovation_star_editor_context():
    """返回当前用户是否可添加创新星内容，以及用于页面展示的身份信息。"""
    user_id = str(session.get('feishu_user_id') or '').strip()
    user_name = _get_current_feishu_user_name()
    department_names = []
    if user_id.startswith('dev_'):
        local_department = str(session.get('user_department') or '').strip()
        if local_department:
            department_names.append(local_department)
    elif user_id and user_name not in INNOVATION_STAR_EDITOR_NAMES:
        try:
            department_rows = permission_manager.get_user_departments(user_id) or []
        except Exception as e:
            _safe_print(f"⚠️ 获取创新星添加权限失败: user={user_name}, error={e}")
            department_rows = []
        for row in department_rows:
            if not isinstance(row, dict):
                continue
            if str(row.get('status') or '').strip() in {'invalid', 'unmapped'}:
                continue
            department_name = str(row.get('name') or '').strip()
            if department_name and department_name not in department_names:
                department_names.append(department_name)
    can_add = bool(user_id) and (user_name in INNOVATION_STAR_EDITOR_NAMES or 'AI部' in department_names)
    return {
        'can_add': can_add,
        'user_id': user_id,
        'user_name': user_name,
        'department_names': department_names,
    }


def _innovation_star_media_names(value):
    """把数据库中的绝对路径或历史相对路径转换成可访问的安全文件名。"""
    names = []
    for raw_path in str(value or '').split(';'):
        clean_path = raw_path.strip()
        if not clean_path:
            continue
        filename = os.path.basename(clean_path.replace('/', os.sep))
        if filename and filename not in names:
            names.append(filename)
    return names


def _innovation_star_items():
    rows = sf_db(
        """
        SELECT ID, TuPian, ShiPin, RIQI, LeiXing, WenAn, TiJiaoRen
        FROM chuangxinxing
        ORDER BY CASE WHEN RIQI IS NULL THEN 1 ELSE 0 END, RIQI DESC, ID DESC
        """
    ) or []
    grouped = {'share': [], 'talk': []}
    for row in rows:
        values = list(row) if isinstance(row, (list, tuple)) else []
        item_id = values[0] if len(values) > 0 else ''
        image_value = values[1] if len(values) > 1 else ''
        video_value = values[2] if len(values) > 2 else ''
        created_at = values[3] if len(values) > 3 else ''
        content_type = str(values[4] if len(values) > 4 and values[4] is not None else '').strip()
        copy_text = str(values[5] if len(values) > 5 and values[5] is not None else '').strip()
        submitter = str(values[6] if len(values) > 6 and values[6] is not None else '').strip()
        type_key = 'talk' if content_type in {'talk', '创新星说', '创新新说'} or '说' in content_type else 'share'
        if isinstance(created_at, datetime):
            created_text = created_at.strftime('%Y-%m-%d %H:%M')
        else:
            created_text = str(created_at or '').strip()
        grouped[type_key].append({
            'id': item_id,
            'images': _innovation_star_media_names(image_value),
            'videos': _innovation_star_media_names(video_value),
            'copy': copy_text,
            'submitter': submitter,
            'created_at': created_text,
        })
    return grouped


def _innovation_star_upload_files(files, media_kind, allowed_extensions):
    valid_files = []
    for uploaded in files:
        if not uploaded or not uploaded.filename:
            continue
        extension = os.path.splitext(uploaded.filename)[1].lower()
        if extension not in allowed_extensions:
            raise ValueError(f'不支持的{media_kind}格式：{extension or "未知格式"}')
        valid_files.append((uploaded, extension))

    if not valid_files:
        return []

    os.makedirs(INNOVATION_STAR_MEDIA_FOLDER, exist_ok=True)
    saved_paths = []
    prefix = 'image' if media_kind == '图片' else 'video'
    try:
        for uploaded, extension in valid_files:
            unique_name = (
                f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_"
                f"{uuid.uuid4().hex[:8]}{extension}"
            )
            full_path = os.path.join(INNOVATION_STAR_MEDIA_FOLDER, unique_name)
            uploaded.save(full_path)
            saved_paths.append(full_path)
    except Exception:
        _innovation_star_cleanup_files(saved_paths)
        raise
    return saved_paths


def _innovation_star_cleanup_files(paths):
    for path in paths:
        try:
            if path and os.path.isfile(path):
                os.remove(path)
        except OSError:
            pass


def _innovation_star_delete_media_files(*path_values):
    """只删除创新星主场目录内、由数据库记录引用的媒体文件。"""
    failed_names = []
    media_names = []
    for path_value in path_values:
        for filename in _innovation_star_media_names(path_value):
            if filename not in media_names:
                media_names.append(filename)
    for filename in media_names:
        full_path = os.path.join(INNOVATION_STAR_MEDIA_FOLDER, filename)
        try:
            if os.path.isfile(full_path):
                os.remove(full_path)
        except OSError:
            failed_names.append(filename)
    return failed_names


def _run_vector_update_async(project_id):
    """后台异步更新单条提案向量与相似度，不阻塞提交接口返回。"""
    def _worker(pid):
        try:
            vector_script = os.path.join(os.path.dirname(__file__), 'build_chuangxin_vectors.py')
            if not os.path.exists(vector_script):
                _safe_print("⚠️ 未找到 build_chuangxin_vectors.py，跳过向量更新")
                return
            env = os.environ.copy()
            env.setdefault('HF_ENDPOINT', 'https://hf-mirror.com')
            cmd = [sys.executable, vector_script, '--bianhao', str(pid)]
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=300, env=env)
            if proc.returncode != 0:
                _safe_print(f"⚠️ 向量增量更新失败({pid}): {proc.stderr}")
            else:
                _safe_print(f"✅ 向量增量更新成功({pid})")
        except Exception as vec_err:
            _safe_print(f"⚠️ 向量更新异常({pid}): {vec_err}")

    t = threading.Thread(target=_worker, args=(project_id,), daemon=True)
    t.start()


# 路由定义
@app.route('/')  # 运行到这立即渲染下方的导航栏网站
def index():  # 定义首页处理函数
    """首页"""
    force_desktop = str(request.args.get('desktop') or '').strip() in {'1', 'true', 'yes'}
    if not force_desktop:
        ua = (request.headers.get('User-Agent') or '').lower()
        mobile_keywords = ['iphone', 'ipad', 'android', 'mobile', 'harmony', 'micromessenger']
        if any(k in ua for k in mobile_keywords):
            return redirect(url_for('innovation.innovation_mobile_index'))
    return render_template('index.html')  # 渲染index.html作为首页


@app.route('/mobile', endpoint='innovation.innovation_mobile_index')
def mobile_index():
    """创新系统移动端首页"""
    return render_template(
        'mobile_index.html',
        feishu_user_name=_get_current_feishu_user_name()
    )


@app.route('/submit', endpoint='innovation.innovation_submit')  # 定义路由,提交页面的路由
def submit():
    """提交页面"""
    from config import DEPARTMENTS  # 从配置文件中获取部门列表
    raw_name = session.get('feishu_user_name', '')
    name_parts = str(raw_name).split('（', 1) if raw_name else []
    feishu_user_name = name_parts[0].strip() if name_parts else str(raw_name).strip()
    return render_template('submit.html', departments=DEPARTMENTS, feishu_user_name=feishu_user_name)  # 渲染网页，对应的部门列表加载上


@app.route('/dashboard')  # 定义路由，积分详情展示页面，已经停用，后续可能会加
def dashboard():
    """仪表板页面"""
    return render_template('dashboard.html')


@app.route('/api/ai/version_update_notice', methods=['POST'])
def api_ai_version_update_notice():
    try:
        data = request.get_json(silent=True) or {}
        template_text = (data.get('template') or '').strip()
        content = (data.get('content') or '').strip()
        department = (data.get('department') or '').strip()
        at_all = bool(data.get('at_all'))
        if not template_text:
            return jsonify({'success': False, 'message': '话术模板不能为空'}), 400
        if not content:
            full_message = template_text
        else:
            full_message = template_text.replace('【在这里填写本次更新的主要内容】', content)
        if not department:
            return jsonify({'success': False, 'message': '推送部门不能为空'}), 400
        message_service = get_message_service()
        result = message_service.send_message_to_department_members(department, full_message, at_all=at_all)
        return jsonify({'success': True, 'message': '发送完成', 'data': result})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/ai/version_update_departments')
def api_ai_version_update_departments():
    try:
        rows = sf_db("SELECT DISTINCT YONGHU FROM feishu_id WHERE FeiShu_ID LIKE 'od_%' ORDER BY YONGHU")
        names = []
        if isinstance(rows, list):
            for r in rows:
                if isinstance(r, dict) and 'YONGHU' in r:
                    names.append(r['YONGHU'])
                elif isinstance(r, str):
                    names.append(r)
        elif isinstance(rows, dict) and 'YONGHU' in rows:
            names.append(rows['YONGHU'])
        elif isinstance(rows, str):
            names.append(rows)
        uniq = []
        seen = set()
        for n in names:
            n = (n or '').strip()
            if n and n not in seen:
                seen.add(n)
                uniq.append({'name': n})
        return jsonify({'success': True, 'data': uniq})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e), 'data': []}), 500


@app.route('/points', endpoint='innovation.innovation_points')  # 用户积分路由，展示积分和对应奖品已经得分记录和消费记录
def points():
    """积分页面"""
    return render_template('points.html', feishu_user_name=_get_current_feishu_user_name())


@app.route('/excellent_cases', endpoint='innovation.excellent_cases')
def excellent_cases():
    """优秀案例页面：展示最高分50分提案全部字段和创新榜单"""
    user_id = session.get('feishu_user_id')
    if not user_id:
        return redirect(url_for('feishu_auth'))

    query_sql = """
        SELECT
            ISNULL(CAST([编号] AS NVARCHAR(100)), '') AS 编号,
            ISNULL(CAST([发起人] AS NVARCHAR(200)), '') AS 发起人,
            ISNULL(CAST([部门] AS NVARCHAR(200)), '') AS 部门,
            ISNULL(CAST([内容] AS NVARCHAR(MAX)), '') AS 内容,
            ISNULL(CAST([解决方案] AS NVARCHAR(MAX)), '') AS 解决方案,
            ISNULL(CAST([发起时间] AS NVARCHAR(100)), '') AS 发起时间
        FROM v_quanyuanchuangxin
        WHERE ISNUMERIC(ISNULL([最高分], 0)) = 1
          AND CAST(ISNULL([最高分], 0) AS INT) = 50
        ORDER BY [发起时间] DESC
    """
    rows = sf_db(query_sql) or []

    cases = []
    for row in rows:
        item = {
            '编号': '' if len(row) < 1 or row[0] is None else str(row[0]),
            '发起人': '' if len(row) < 2 or row[1] is None else str(row[1]),
            '部门': '' if len(row) < 3 or row[2] is None else str(row[2]),
            '内容': '' if len(row) < 4 or row[3] is None else str(row[3]),
            '解决方案': '' if len(row) < 5 or row[4] is None else str(row[4]),
            '发起时间': '' if len(row) < 6 or row[5] is None else str(row[5]),
        }
        # 按前端固定字段映射：题号=编号，姓名=发起人，部门=部门，内容=内容，详情=解决方案
        item['_display_no'] = item.get('编号') or '未知编号'
        item['_display_name'] = item.get('发起人') or '未知'
        item['_display_dept'] = item.get('部门') or '未知部门'
        item['_display_content'] = item.get('内容') or '未填写'
        item['_display_detail'] = item.get('解决方案') or '未填写'
        item['_display_time'] = item.get('发起时间') or ''
        cases.append(item)

    leaderboard_sql = """
        SELECT TOP 11
            SUM(CASE WHEN ISNUMERIC(ISNULL([最高分], 0)) = 1 THEN CAST(ISNULL([最高分], 0) AS INT) ELSE 0 END) AS 总分,
            ISNULL(CAST([发起人] AS NVARCHAR(200)), '') AS 发起人
        FROM v_QuanYuanChuangXin
        WHERE [发起时间] >= '2025-07-01'
          AND [发起时间] < '2026-06-01'
        GROUP BY [发起人]
        HAVING SUM(CASE WHEN ISNUMERIC(ISNULL([最高分], 0)) = 1 THEN CAST(ISNULL([最高分], 0) AS INT) ELSE 0 END) > 0
        ORDER BY SUM(CASE WHEN ISNUMERIC(ISNULL([最高分], 0)) = 1 THEN CAST(ISNULL([最高分], 0) AS INT) ELSE 0 END) DESC
    """
    leaderboard_rows = sf_db(leaderboard_sql) or []
    leaderboard = []
    for idx, row in enumerate(leaderboard_rows, start=1):
        if isinstance(row, dict):
            score = row.get('总分') or row.get('total_score') or row.get('SUM') or 0
            name = row.get('发起人') or row.get('name') or ''
        else:
            values = list(row) if isinstance(row, (list, tuple)) else []
            score = values[0] if len(values) > 0 else 0
            name = values[1] if len(values) > 1 else ''
        leaderboard.append({
            'rank': idx,
            'name': str(name or '').strip() or '未知',
            'score': int(float(score or 0)) if str(score or '').strip() else 0,
        })

    try:
        star_items = _innovation_star_items()
    except Exception as e:
        _safe_print(f"⚠️ 获取创新星享/创新星说内容失败: {e}")
        star_items = {'share': [], 'talk': []}
    editor_context = _innovation_star_editor_context()

    return render_template(
        'excellent_cases.html',
        cases=cases,
        total_count=len(cases),
        leaderboard=leaderboard,
        star_share_items=star_items['share'],
        star_talk_items=star_items['talk'],
        can_add_star_content=editor_context['can_add'],
        star_editor_name=editor_context['user_name'],
    )


@app.route('/api/innovation_star_content', methods=['POST'])
def add_innovation_star_content():
    if not session.get('feishu_user_id'):
        return jsonify({'success': False, 'message': '请先登录后再操作'}), 401

    editor_context = _innovation_star_editor_context()
    if not editor_context['can_add']:
        return jsonify({'success': False, 'message': '当前账号没有添加权限'}), 403

    type_key = str(request.form.get('content_type') or 'share').strip().lower()
    type_labels = {'share': '创新新享', 'talk': '创新新说'}
    if type_key not in type_labels:
        return jsonify({'success': False, 'message': '内容分类无效'}), 400

    copy_text = str(request.form.get('copy') or '').strip()
    image_paths = []
    video_paths = []
    try:
        image_paths = _innovation_star_upload_files(
            request.files.getlist('images'),
            '图片',
            INNOVATION_STAR_IMAGE_EXTENSIONS,
        )
        video_paths = _innovation_star_upload_files(
            request.files.getlist('videos'),
            '视频',
            INNOVATION_STAR_VIDEO_EXTENSIONS,
        )
    except ValueError as e:
        _innovation_star_cleanup_files(image_paths + video_paths)
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        _innovation_star_cleanup_files(image_paths + video_paths)
        _safe_print(f"❌ 保存创新星媒体失败: {e}")
        return jsonify({'success': False, 'message': '文件保存失败，请稍后重试'}), 500

    def _sql_nullable(value):
        clean_value = str(value or '').strip()
        if not clean_value:
            return 'NULL'
        return f"N'{_escape_sql_literal_for_pytds(clean_value)}'"

    try:
        image_value = ';'.join(image_paths)
        video_value = ';'.join(video_paths)
        insert_sql = f"""
            INSERT INTO chuangxinxing (TuPian, ShiPin, RIQI, LeiXing, WenAn, TiJiaoRen)
            VALUES (
                {_sql_nullable(image_value)},
                {_sql_nullable(video_value)},
                GETDATE(),
                {_sql_nullable(type_labels[type_key])},
                {_sql_nullable(copy_text)},
                {_sql_nullable(editor_context['user_name'])}
            )
        """
        dui_db(insert_sql)
    except Exception as e:
        _innovation_star_cleanup_files(image_paths + video_paths)
        _safe_print(f"❌ 写入创新星内容失败: {e}")
        return jsonify({'success': False, 'message': '内容保存失败，请稍后重试'}), 500

    return jsonify({
        'success': True,
        'message': '添加成功',
        'section': type_key,
    })


@app.route('/api/innovation_star_content/<int:item_id>', methods=['DELETE'])
def delete_innovation_star_content(item_id):
    if not session.get('feishu_user_id'):
        return jsonify({'success': False, 'message': '请先登录后再操作'}), 401

    editor_context = _innovation_star_editor_context()
    if not editor_context['can_add']:
        return jsonify({'success': False, 'message': '当前账号没有删除权限'}), 403

    try:
        rows = sf_db(
            f"""
            SELECT TOP 1 TuPian, ShiPin, LeiXing
            FROM chuangxinxing
            WHERE ID = {int(item_id)}
            """
        ) or []
    except Exception as e:
        _safe_print(f"❌ 查询待删除创新星内容失败: id={item_id}, error={e}")
        return jsonify({'success': False, 'message': '内容查询失败，请稍后重试'}), 500

    if not rows:
        return jsonify({'success': False, 'message': '内容不存在或已被删除'}), 404

    first = rows[0]
    values = list(first) if isinstance(first, (list, tuple)) else []
    image_value = values[0] if len(values) > 0 else ''
    video_value = values[1] if len(values) > 1 else ''
    content_type = str(values[2] if len(values) > 2 and values[2] is not None else '').strip()
    section = 'talk' if content_type in {'talk', '创新星说', '创新新说'} or '说' in content_type else 'share'

    try:
        dui_db(f"DELETE FROM chuangxinxing WHERE ID = {int(item_id)}")
    except Exception as e:
        _safe_print(f"❌ 删除创新星内容失败: id={item_id}, error={e}")
        return jsonify({'success': False, 'message': '删除失败，请稍后重试'}), 500

    failed_files = _innovation_star_delete_media_files(image_value, video_value)
    return jsonify({
        'success': True,
        'message': '删除成功' if not failed_files else '内容已删除，部分媒体文件清理失败',
        'section': section,
        'file_cleanup_complete': not failed_files,
    })


@app.route('/manage', endpoint='innovation.innovation_manage')
def manage():
    """管理页面 - 所有用户可访问"""
    force_desktop = str(request.args.get('desktop') or '').strip() in {'1', 'true', 'yes'}
    if (not force_desktop) and _is_mobile_request():
        return redirect(url_for('innovation.innovation_manage_mobile'))

    username = session.get('username', '')  # 获取会话中的用户名

    raw_name = session.get('feishu_user_name', '')
    name_parts = str(raw_name).split('（', 1) if raw_name else []
    feishu_user_name = name_parts[0].strip() if name_parts else str(raw_name).strip()
    feishu_user_id = session.get('feishu_user_id', '')

    # 定义委员会成员名单
    committee_members = ['周俊成', '陶晓飞', '孙军', '毕景春', '李昌瀚', '蔡晶', '韩雅俊','孙洁','陈子烨']

    # 检查当前用户是否为委员会成员
    is_committee_member = feishu_user_name in committee_members

    _safe_print(f"🔍 创新管理页面访问 - 用户: {feishu_user_name}, 是否委员会成员: {is_committee_member}")

    return render_template('manage.html',
                           is_admin=True,
                           feishu_user_name=feishu_user_name,
                           feishu_user_id=feishu_user_id,
                           is_committee_member=is_committee_member)  # 渲染创新管理页面，传递用户信息和权限


@app.route('/manage/mobile', endpoint='innovation.innovation_manage_mobile')
def manage_mobile():
    """创新承接管理移动端页面"""
    raw_name = session.get('feishu_user_name', '')
    name_parts = str(raw_name).split('（', 1) if raw_name else []
    feishu_user_name = name_parts[0].strip() if name_parts else str(raw_name).strip()
    return render_template('manage_mobile.html', feishu_user_name=feishu_user_name)

# ===== 创新点赞/评论 API =====
@app.route('/api/innovation/react', methods=['POST'])
def innovation_react():
    """对创新编号进行点赞/评论：
    - 点赞：仍按编号+当前用户唯一记录，存在则只更新点赞标记；
    - 评论：每次提交都会新增一条评论记录，允许同一用户多次评论。
    """
    try:
        uploaded_image_path = None
        data = None
        if request.files or (request.content_type and request.content_type.startswith('multipart/form-data')):
            raw_id = request.form.get('id', '')
            like = request.form.get('like')
            if isinstance(like, str):
                like_norm = like.strip().lower()
                if like_norm in {'true', '1', 'yes', 'y', 'on'}:
                    like = True
                elif like_norm in {'false', '0', 'no', 'n', 'off'}:
                    like = False
                else:
                    like = None
            comment = request.form.get('comment')

            img = request.files.get('image')
            if img and img.filename:
                filename = secure_filename(img.filename)
                ext = os.path.splitext(filename)[1].lower()
                allowed_image_exts = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp'}
                if ext and ext not in allowed_image_exts:
                    return jsonify({'success': False, 'message': f'不支持的图片格式: {ext}'}), 400
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
                filename = f"{timestamp}_{filename}"
                save_path = os.path.join(LIKE_IMAGE_FOLDER, filename)
                img.save(save_path)
                uploaded_image_path = save_path.replace('\\', '/')
        else:
            data = request.get_json() or {}
            raw_id = data.get('id', '')
            like = data.get('like')
            comment = data.get('comment')

        m = re.search(r"(\d+)", str(raw_id))
        project_id = int(m.group(1)) if m else 0

        if not project_id:
            if uploaded_image_path:
                try:
                    os.remove(uploaded_image_path.replace('/', '\\'))
                except Exception:
                    pass
            return jsonify({'success': False, 'message': '缺少编号'}), 400

        user_name = _get_current_feishu_user_name()
        if not user_name:
            if uploaded_image_path:
                try:
                    os.remove(uploaded_image_path.replace('/', '\\'))
                except Exception:
                    pass
            return jsonify({'success': False, 'message': '未识别飞书用户'}), 401

        user_esc = user_name.replace("'", "''")
        comment_esc = (comment if comment is not None else '').replace("'", "''")
        image_esc = (uploaded_image_path or '').replace("'", "''")
        like_is_set = like is not None
        like_val = 1 if like_is_set and like else 0

        sql_list = []

        def _scalar_first(v):
            if v is None:
                return None
            if isinstance(v, (list, tuple)):
                return v[0] if v else None
            return v

        def _has_column(table_name, column_name):
            try:
                sql = (
                    "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
                    f"WHERE TABLE_NAME='{table_name}' AND COLUMN_NAME=N'{column_name}'"
                )
                r = sf_db(sql) or []
                if not r:
                    return False
                c = _scalar_first(r[0])
                return int(c) > 0
            except Exception:
                return False

        has_image_column = _has_column('chuangxin_dianzan', '图片')
        if uploaded_image_path and (not has_image_column):
            try:
                os.remove(uploaded_image_path.replace('/', '\\'))
            except Exception:
                pass
            return jsonify({'success': False, 'message': '数据库缺少“图片”字段，无法保存图片'}), 500

        # 1) 处理点赞：保持每个用户对同一编号只有一条“点赞状态”记录（评论为空或空白）
        if like_is_set:
            like_where = (
                f"编号 = {project_id} AND 操作人 = '{user_esc}' "
                "AND (评论 IS NULL OR LTRIM(RTRIM(评论)) = '')"
            )
            exist_like = sf_db(
                f"SELECT COUNT(*) FROM chuangxin_dianzan WHERE {like_where}",
                single=True
            ) or 0
            if exist_like > 0:
                sql_list.append(
                    f"UPDATE chuangxin_dianzan "
                    f"SET 点赞 = {like_val}, 操作时间 = GETDATE() "
                    f"WHERE {like_where}"
                )
            else:
                sql_list.append(
                    f"INSERT INTO chuangxin_dianzan (编号, 操作人, 点赞, 评论, 操作时间) "
                    f"VALUES ({project_id}, '{user_esc}', {like_val}, NULL, GETDATE())"
                )

        # 2) 处理评论：每次提交都插入一条新的评论记录，允许多次评论
        has_comment = (comment is not None and comment_esc.strip() != '')
        has_image = bool(uploaded_image_path)
        # #region debug-point A:react-branch
        _innovation_debug_report('A', 'innovation.web_app:innovation_react:branch', 'innovation react resolved comment branch state', {'project_id': project_id, 'user_name': user_name, 'has_comment': has_comment, 'has_image': has_image, 'like_is_set': bool(like_is_set), 'content_type': str(request.content_type or '')})
        # #endregion
        if has_comment or has_image:
            if has_image_column:
                sql_list.append(
                    f"INSERT INTO chuangxin_dianzan (编号, 操作人, 点赞, 评论, 图片, 操作时间) "
                    f"VALUES ({project_id}, '{user_esc}', 0, '{comment_esc}', '{image_esc}', GETDATE())"
                )
            else:
                sql_list.append(
                    f"INSERT INTO chuangxin_dianzan (编号, 操作人, 点赞, 评论, 操作时间) "
                    f"VALUES ({project_id}, '{user_esc}', 0, '{comment_esc}', GETDATE())"
                )

        if not sql_list:
            if uploaded_image_path:
                try:
                    os.remove(uploaded_image_path.replace('/', '\\'))
                except Exception:
                    pass
            return jsonify({'success': False, 'message': '缺少点赞或评论内容'}), 400

        for sql in sql_list:
            _safe_print('[innovation_react]', 'id=', project_id, 'sql=', sql)
            dui_db(sql)

        notify_result = None
        if has_comment or has_image:
            notify_result = _innovation_notify_initiator_on_comment(
                project_id,
                user_name,
                comment_text=comment or '',
                has_image=has_image
            )
        # #region debug-point A:react-notify-result
        _innovation_debug_report('A', 'innovation.web_app:innovation_react:notify_result', 'innovation react finished notify branch', {'project_id': project_id, 'user_name': user_name, 'notify_result': notify_result or {}})
        # #endregion

        return jsonify({'success': True, 'notify': notify_result})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/innovation/react/comment', methods=['PUT'])
def innovation_update_comment():
    """修改评论（按编号+当前用户唯一记录进行修改）"""
    try:
        data = request.get_json() or {}
        raw_id = data.get('id', '')
        m = re.search(r"(\d+)", str(raw_id))
        project_id = int(m.group(1)) if m else 0
        comment = data.get('comment', '')

        if not project_id:
            return jsonify({'success': False, 'message': '缺少编号'}), 400

        user_name = _get_current_feishu_user_name()
        if not user_name:
            return jsonify({'success': False, 'message': '未识别飞书用户'}), 401

        user_esc = user_name.replace("'", "''")
        comment_esc = (comment or '').replace("'", "''")

        exist_sql = f"SELECT COUNT(*) FROM chuangxin_dianzan WHERE 编号 = {project_id} AND 操作人 = '{user_esc}'"
        exist_count = sf_db(exist_sql, single=True) or 0
        # #region debug-point D:update-comment-entry
        _innovation_debug_report('D', 'innovation.web_app:innovation_update_comment:entry', 'innovation update comment called', {'project_id': project_id, 'user_name': user_name, 'exist_count': int(exist_count or 0), 'has_comment': bool(str(comment or '').strip())})
        # #endregion

        if exist_count == 0:
            insert_sql = f"""
                INSERT INTO chuangxin_dianzan (编号, 操作人, 点赞, 评论, 操作时间)
                VALUES ({project_id}, '{user_esc}', 0, '{comment_esc}', GETDATE())
            """
            _safe_print('[innovation_update_comment][insert]', 'id=', project_id, 'comment=', comment_esc)
            dui_db(insert_sql)
            notify_result = None
            if str(comment or '').strip():
                notify_result = _innovation_notify_initiator_on_comment(
                    project_id,
                    user_name,
                    comment_text=comment or '',
                    has_image=False
                )
            # #region debug-point D:update-comment-notify-result
            _innovation_debug_report('D', 'innovation.web_app:innovation_update_comment:notify_result', 'innovation update comment finished notify branch', {'project_id': project_id, 'user_name': user_name, 'notify_result': notify_result or {}, 'created': True})
            # #endregion
            return jsonify({'success': True, 'created': True, 'notify': notify_result})

        update_sql = f"""
            UPDATE chuangxin_dianzan
            SET 评论 = '{comment_esc}', 操作时间 = GETDATE()
            WHERE 编号 = {project_id} AND 操作人 = '{user_esc}'
        """
        _safe_print('[innovation_update_comment][update]', 'id=', project_id, 'comment=', comment_esc)
        dui_db(update_sql)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/innovation/react/stats', methods=['GET'])
def innovation_get_react_stats():
    try:
        raw_id = request.args.get('id', '')
        # 兼容形如 "#648" 或包含非数字字符的编号
        m = re.search(r"(\d+)", str(raw_id))
        project_id = int(m.group(1)) if m else 0
        if not project_id:
            return jsonify({'success': False, 'message': '缺少编号'}), 400
        likes = sf_db(f"SELECT COUNT(*) FROM chuangxin_dianzan WHERE 编号 = {project_id} AND 点赞 = 1", single=True) or 0
        like_rows = sf_db(f"SELECT 操作人 FROM chuangxin_dianzan WHERE 编号 = {project_id} AND 点赞 = 1 ORDER BY 操作时间 DESC") or []
        like_users = []
        for r in like_rows:
            if isinstance(r, (list, tuple)):
                like_users.append(r[0] or '')
            else:
                like_users.append(r or '')
        def _scalar_first(v):
            if v is None:
                return None
            if isinstance(v, (list, tuple)):
                return v[0] if v else None
            return v

        def _has_column(table_name, column_name):
            try:
                sql = (
                    "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
                    f"WHERE TABLE_NAME='{table_name}' AND COLUMN_NAME=N'{column_name}'"
                )
                r = sf_db(sql) or []
                if not r:
                    return False
                c = _scalar_first(r[0])
                return int(c) > 0
            except Exception:
                return False

        has_image_column = _has_column('chuangxin_dianzan', '图片')
        if has_image_column:
            comments_rows = sf_db(
                f"SELECT 操作人, 评论, 操作时间, 图片 "
                f"FROM chuangxin_dianzan "
                f"WHERE 编号 = {project_id} AND ("
                f"(评论 IS NOT NULL AND LTRIM(RTRIM(评论)) <> '') OR (图片 IS NOT NULL AND LTRIM(RTRIM(图片)) <> '')"
                f") "
                f"ORDER BY 操作时间 DESC"
            ) or []
        else:
            comments_rows = sf_db(
                f"SELECT 操作人, 评论, 操作时间 "
                f"FROM chuangxin_dianzan "
                f"WHERE 编号 = {project_id} AND 评论 IS NOT NULL AND LTRIM(RTRIM(评论)) <> '' "
                f"ORDER BY 操作时间 DESC"
            ) or []
        _safe_print('[innovation_get_react_stats]', 'id=', project_id, 'rows_count=', len(comments_rows))
        def _fmt_time(t):
            try:
                return t.strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                return str(t or '')
        def _fmt_image_url(p):
            if not p:
                return None
            if isinstance(p, bytes):
                try:
                    p = p.decode('utf-8', errors='ignore')
                except Exception:
                    p = str(p)
            p = str(p).strip().strip('"\'' )
            if not p:
                return None
            base = p.replace('\\', '/').split('/')[-1].strip()
            if not base:
                return None
            return f"/innovation/like_images/{quote(base)}"
        comments = []
        for r in comments_rows:
            if isinstance(r, dict):
                comments.append({
                    'operator': (r.get('操作人', '') or '').strip(),
                    'comment': (r.get('评论', '') or '').strip(),
                    'time': _fmt_time(r.get('操作时间')),
                    'image_url': _fmt_image_url(r.get('图片')) if has_image_column else None
                })
            elif isinstance(r, (list, tuple)):
                comments.append({
                    'operator': (r[0] or '').strip(),
                    'comment': (r[1] or '').strip(),
                    'time': _fmt_time(r[2]),
                    'image_url': _fmt_image_url(r[3]) if (has_image_column and len(r) > 3) else None
                })
            else:
                # 未知结构，尝试直接字符串化
                try:
                    comments.append({'operator': '', 'comment': str(r or ''), 'time': '', 'image_url': None})
                except Exception:
                    pass
        _safe_print('[innovation_get_react_stats]', 'id=', project_id, 'comments=', comments)
        user_name = session.get('feishu_user_name', '')
        mine_like = None
        if user_name:
            mine_like = sf_db(f"SELECT 点赞 FROM chuangxin_dianzan WHERE 编号 = {project_id} AND 操作人 = '{user_name.replace("'", "''")}'", single=True)
            mine_like = bool(mine_like) if mine_like is not None else False
        return jsonify({'success': True, 'likes': int(likes), 'like_users': like_users, 'comments': comments, 'mine_like': mine_like})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/submit_innovation', methods=['POST'])  # 定义新项目提交路由
def submit_innovation():
    try:
        message_service = get_message_service()  # 获取消息服务实例（根据URL参数选择公司配置）
        title = request.form.get('title', '').strip()  # 获取表单数据，移除空格
        content = request.form.get('content', '').strip()  # 获取表单内容，移除空格
        raw_name = session.get('feishu_user_name', '')
        name_parts = str(raw_name).split('（', 1) if raw_name else []
        feishu_user_name = name_parts[0].strip() if name_parts else str(raw_name).strip()
        form_initiator = request.form.get('initiator', '').strip()
        if feishu_user_name == '张雯' and form_initiator:
            initiator = form_initiator
        else:
            initiator = str(feishu_user_name).strip() if feishu_user_name else form_initiator
        deadline = request.form.get('deadline', '').strip()  # 获取表单截止日期，移除空格
        departments = request.form.get('department', '').strip()  # 获取表单承接人，移除空格
        category = request.form.get('category', '').strip()  # 获取表单创新类别，移除空格
        solution = request.form.get('solution', '').strip()  # 获取表单解决方案，移除空格

        if not title or not content or not initiator or not category or not solution:  # 标题、内容、发起人、创新类别和解决方案只要有一个为空直接报错，返回400
            return jsonify({
                'success': False,
                'message': '标题、内容、发起人、创新类别和解决方案不能为空'
            }), 400

        image_paths = []  # 处理多张图片上传
        if 'images' in request.files:  # 若是request中存在images
            files = request.files.getlist('images')  # 获取所有上传的图片文件
            for file in files:
                if file and file.filename:  # 确保文件存在且文件名存在
                    filename = secure_filename(file.filename)  # 使用secure_filename ,防止姓名冲突
                    timestamp = datetime.now().strftime(
                        '%Y%m%d_%H%M%S_%f')  # 将现在的时间设定为固定格式，包含微秒避免重名
                    filename = f"{timestamp}_{filename}"  # 文件名称为时间序列拼接上一个文件名
                    image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)  # 构造文件保存路径
                    file.save(image_path)  # 保存图片
                    image_paths.append(_database_upload_path(filename))  # 数据库存相对路径

        video_paths = []
        if 'videos' in request.files:
            video_files = request.files.getlist('videos')
            allowed_video_exts = {'.mp4', '.webm', '.ogg', '.mov'}
            for file in video_files:
                if file and file.filename:
                    filename = secure_filename(file.filename)
                    ext = os.path.splitext(filename)[1].lower()
                    if ext and ext not in allowed_video_exts:
                        return jsonify({'success': False, 'message': f'不支持的视频格式: {ext}'}), 400
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
                    filename = f"{timestamp}_{filename}"
                    video_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    file.save(video_path)
                    video_paths.append(_database_upload_path(filename))

        # 将多个图片路径用分号分隔存储
        image_paths_str = ';'.join(image_paths) if image_paths else ''
        video_paths_str = ';'.join(video_paths) if video_paths else ''

        title_escaped = _escape_sql_literal_for_pytds(title)
        content_escaped = _escape_sql_literal_for_pytds(content)
        initiator_escaped = _escape_sql_literal_for_pytds(initiator)
        category_escaped = _escape_sql_literal_for_pytds(category)
        solution_escaped = _escape_sql_literal_for_pytds(solution) if solution else ''
        deadline_escaped = _escape_sql_literal_for_pytds(deadline) if deadline else ''
        image_paths_str = _escape_sql_literal_for_pytds(image_paths_str)
        video_paths_str = _escape_sql_literal_for_pytds(video_paths_str)

        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 最新的时间按照%Y-%m-%d %H:%M:%S格式输出

        def _scalar_first(v):
            if v is None:
                return None
            if isinstance(v, (list, tuple)):
                return v[0] if v else None
            return v

        def _has_column(table_name, column_name):
            try:
                sql = (
                    "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
                    f"WHERE TABLE_NAME='{table_name}' AND COLUMN_NAME=N'{column_name}'"
                )
                r = sf_db(sql) or []
                if not r:
                    return False
                c = _scalar_first(r[0])
                return int(c) > 0
            except Exception:
                return False

        has_video_column = _has_column('chuangxin_tibao1', '视频详情')
        _ensure_nvarchar_max_column('chuangxin_tibao1', '解决方案')
        _ensure_nvarchar_max_column('chuangxin_tibao1', '内容')
        if (not has_video_column) and video_paths_str:
            try:
                dui_db("ALTER TABLE chuangxin_tibao1 ADD 视频详情 NVARCHAR(MAX) NULL")
                has_video_column = _has_column('chuangxin_tibao1', '视频详情')
            except Exception:
                has_video_column = False
        if video_paths_str and (not has_video_column):
            for p in video_paths:
                try:
                    os.remove(p)
                except Exception:
                    pass
            return jsonify({'success': False, 'message': '数据库缺少“视频详情”字段，无法保存视频，请联系管理员处理'}), 500

        if has_video_column:
            insert_sql = f"""
                        INSERT INTO chuangxin_tibao1 (标题, 内容, 发起人, 发起时间, 截止时间, 得分, 状态, 创新类别, 解决方案, 图片详情, 视频详情)
                        VALUES ('{title_escaped}', '{content_escaped}', '{initiator_escaped}', '{current_time}', '{deadline_escaped}', 0, '待承接', '{category_escaped}', '{solution_escaped}', '{image_paths_str}', '{video_paths_str}')
                    """
        else:
            insert_sql = f"""
                        INSERT INTO chuangxin_tibao1 (标题, 内容, 发起人, 发起时间, 截止时间, 得分, 状态, 创新类别, 解决方案, 图片详情)
                        VALUES ('{title_escaped}', '{content_escaped}', '{initiator_escaped}', '{current_time}', '{deadline_escaped}', 0, '待承接', '{category_escaped}', '{solution_escaped}', '{image_paths_str}')
                    """  # 插入提交信息，包括标题，内容，发起人,发起时间，截止时间,得分，状态，创新类别，解决方案，图片详情

        try:  # 尝试对数据库进行操作，执行插入操作
            dui_db(insert_sql)
        except Exception as insert_error:
            raise Exception(f"数据插入失败: {insert_error}")  # 否则插入失败

        try:  # 尝试获取插入id
            scope_id_sql = "SELECT SCOPE_IDENTITY() as new_id"
            scope_result = sf_db(scope_id_sql)  # 执行数据库操作
            if scope_result and len(scope_result) > 0 and scope_result[
                0] is not None:  # 如果查询是否生效，若是scope_result不为空且存在、长度大于0，则执行生效
                project_id = int(scope_result[0])  # 项目的id就被赋值为scope_result
                _safe_print(f"✅ 通过SCOPE_IDENTITY获取到项目ID: {project_id}")
            else:
                raise Exception("SCOPE_IDENTITY返回空值")
        except Exception as scope_error:
            _safe_print(f"⚠️ SCOPE_IDENTITY方法失败: {scope_error}，尝试备用方法")

            get_id_sql = f"""
                SELECT TOP 1 编号 FROM chuangxin_tibao1 
                WHERE 发起人 = '{initiator_escaped}' 
                AND 标题 = '{title_escaped}' 
                ORDER BY 编号 DESC
            """  # 方法2：查询最新插入的记录

            _safe_print(f"🔍 执行查询SQL: {get_id_sql}")

            try:
                project_id_result = sf_db(get_id_sql)
                _safe_print(f"📊 查询结果: {project_id_result}")

                if not project_id_result or len(project_id_result) == 0:
                    raise Exception(f"无法获取插入的项目ID。查询条件：发起人={initiator_escaped}, 标题={title_escaped}")

                project_id = int(project_id_result[0])
                _safe_print(f"✅ 通过查询获取到项目ID: {project_id}")
            except Exception as query_error:
                _safe_print(f"❌ 查询项目ID失败: {query_error}")
                raise Exception(f"查询项目ID失败: {query_error}")

        dept_list = []
        if departments:
            dept_list = [dept.strip() for dept in departments.split(',') if dept.strip()]
            _safe_print(f"👥 承接部门列表: {dept_list}")
        else:
            _safe_print(f"⚠️ 未接收到承接部门数据，departments={departments}")

        # 为每个承接人插入流转记录
        # 在submit_innovation函数中添加调试信息
        for dept in dept_list:
            _safe_print(f"🔍 准备插入承接人: '{dept}', 长度: {len(dept)}")
            dept_escaped = dept.replace("'", "''")
            flow_sql = f"""
                INSERT INTO chuangxin_liuzhuan1 (项目编号, 发起人, 承接人, 流转次数, 流转时间, 状态)
                VALUES ({project_id}, '{initiator_escaped}', '{dept_escaped}', 1, '{current_time}', '待承接')
            """
            _safe_print(f"🔍 SQL语句: {flow_sql}")

            try:
                dui_db(flow_sql)
                _safe_print(f"✅ 为 {dept} 创建流转记录成功")
            except Exception as flow_error:
                _safe_print(f"❌ 为 {dept} 创建流转记录失败: {flow_error}")
        # 5. 发送飞书通知给承接部门的所有成员
        notify_summary = {
            'success': 0,
            'failed': 0,
            'failed_departments': []
        }
        try:
            # 构建通知消息模板
            manage_url = "http://223.78.73.100:8000/innovation/manage"
            deadline_text = f"\n截止时间：{deadline}" if deadline else ""
            watcher_ids, initiator_departments, matched_operation_departments = _innovation_collect_operation_watcher_ids(
                initiator,
                dept_list
            )
            _safe_print(
                f"📌 发起人部门: {initiator_departments or []}, "
                f"命中的运营部门: {matched_operation_departments or []}, "
                f"额外观察人数量: {len(watcher_ids)}"
            )

            notification_message = f"""🚀 新的创新项目待承接

项目标题：{title}
承接部门：{', '.join(dept_list)}
发起人：{initiator}{deadline_text}

项目内容：
{content}

📋 点击链接进入管理页面处理：
{manage_url}"""

            # 先按部门收集成员，再跨部门去重后发送，避免同一承接人收到多条重复通知
            total_success = 0
            total_failed = 0
            unique_member_ids = []
            unique_member_seen = set()

            for dept_name in dept_list:
                try:
                    dept_members = message_service.get_department_contacts(dept_name) or []
                    dept_unique_count = 0
                    for member_id in dept_members:
                        oid = str(member_id or '').strip()
                        if not oid or oid in unique_member_seen:
                            continue
                        unique_member_seen.add(oid)
                        unique_member_ids.append(oid)
                        dept_unique_count += 1
                    if dept_unique_count <= 0:
                        notify_summary['failed_departments'].append(dept_name)
                    _safe_print(f"📋 部门 {dept_name} 去重后新增通知对象 {dept_unique_count} 人")
                except Exception as dept_error:
                    _safe_print(f"⚠️ 读取部门 {dept_name} 成员失败: {dept_error}")
                    notify_summary['failed_departments'].append(dept_name)

            for watcher_id in watcher_ids:
                if not watcher_id or watcher_id in unique_member_seen:
                    continue
                unique_member_seen.add(watcher_id)
                unique_member_ids.append(watcher_id)
            if watcher_ids:
                _safe_print(f"📋 已追加运营观察人通知对象 {len(watcher_ids)} 个")

            for member_id in unique_member_ids:
                try:
                    ok = bool(message_service.send_message(member_id, notification_message))
                    if ok:
                        total_success += 1
                        notify_summary['success'] += 1
                    else:
                        total_failed += 1
                        notify_summary['failed'] += 1
                        _safe_print(f"❌ 承接人通知发送失败: {member_id}")
                except Exception as member_error:
                    total_failed += 1
                    notify_summary['failed'] += 1
                    _safe_print(f"⚠️ 承接人通知发送异常 {member_id}: {member_error}")

            _safe_print(
                f"📱 创新项目通知发送完成: 去重后接收人 {len(unique_member_ids)} 个，"
                f"成功 {total_success} 条，失败 {total_failed} 条"
            )

        except Exception as feishu_error:
            _safe_print(f"⚠️ 飞书通知发送失败: {feishu_error}")
            notify_summary['failed'] += 1
            notify_summary['failed_departments'].extend(dept_list)

        # 6. 提案提交后，后台异步更新向量与相似度（按编号）
        _run_vector_update_async(project_id)

        _safe_print(f"🎉 创新项目提交完成，项目ID: {project_id}")

        return jsonify({
            'success': True,
            'message': '创新项目提交成功！',
            'project_id': project_id,
            'notify': notify_summary
        })

    except Exception as e:
        _safe_print(f"💥 提交失败详细错误: {e}")
        _safe_print(f"📍 错误类型: {type(e).__name__}")
        import traceback
        _safe_print(f"📋 错误堆栈: {traceback.format_exc()}")

        return jsonify({
            'success': False,
            'message': f'提交失败: {str(e)}'
        }), 500


@app.route('/api/get_innovations', methods=['GET'])
def get_innovations():
    """获取创新项目列表API"""
    try:
        # 获取分页参数
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 10))

        # 获取筛选参数
        status_filter = request.args.get('status', '').strip()
        department_filter = request.args.get('department', '').strip()
        initiator_department_filter = request.args.get('initiator_department', '').strip()
        operation_type_filter = request.args.get('operation_type', '').strip()
        start_date_filter = request.args.get('start_date', '').strip()
        end_date_filter = request.args.get('end_date', '').strip()
        search_query = request.args.get('search', '').strip()
        max_score_filters = request.args.getlist('max_score') or []
        favorite_only = str(request.args.get('favorite_only') or '').strip().lower() in {'1', 'true', 'yes'}
        favorite_category = str(request.args.get('favorite_category') or '').strip()
        if favorite_category not in {'长期看效果', '暂未有解决方案'}:
            favorite_category = ''
        current_user_name = _get_current_feishu_user_name()

        # 计算偏移量
        offset = (page - 1) * per_page

        # 构建WHERE条件
        where_conditions = []

        # 部门筛选：直接在SQL中进行
        if department_filter:
            where_conditions.append(f"""编号 IN (
                SELECT DISTINCT 项目编号 
                FROM chuangxin_liuzhuan1 
                WHERE 承接人 = '{department_filter.replace("'", "''")}'
            )""")

        # 发起部门筛选：通过v_quanyuanchuangxin视图进行
        if initiator_department_filter:
            where_conditions.append(f"""编号 IN (
                SELECT DISTINCT 编号 
                FROM v_quanyuanchuangxin 
                WHERE 部门 = '{initiator_department_filter.replace("'", "''")}'
            )""")

        # 操作类型筛选：通过chuangxin_liuzhuan1表进行
        if operation_type_filter:
            if operation_type_filter == 'adopted':
                # 采纳：包括'采纳，立即执行'和'采纳，暂缓执行'
                where_conditions.append(f"""编号 IN (
                    SELECT DISTINCT 项目编号 
                    FROM chuangxin_liuzhuan1 
                    WHERE 操作类型 IN ('采纳，立即执行', '采纳，暂缓执行')
                )""")
            elif operation_type_filter == 'not_adopted':
                # 未采纳：包括'不采纳，不执行'和'重复提案'
                where_conditions.append(f"""编号 IN (
                    SELECT DISTINCT 项目编号 
                    FROM chuangxin_liuzhuan1 
                    WHERE 操作类型 IN ('不采纳，不执行', '重复提案')
                )""")
            elif operation_type_filter == 'pending':
                # 待处理
                where_conditions.append(f"""编号 IN (
                    SELECT DISTINCT 项目编号 
                    FROM chuangxin_liuzhuan1 
                    WHERE 操作类型 = '待处理'
                )""")
            elif operation_type_filter == 'disputed':
                # 有异议，上会
                where_conditions.append(f"""编号 IN (
                    SELECT DISTINCT 项目编号 
                    FROM chuangxin_liuzhuan1 
                    WHERE 操作类型 = '有异议，上会'
                )""")
            elif operation_type_filter == 'disputed_10':
                # 有异议，上会且v_quanyuanchuangxin最高分为10分
                where_conditions.append(f"""编号 IN (
                    SELECT DISTINCT l.项目编号
                    FROM chuangxin_liuzhuan1
                    l INNER JOIN v_quanyuanchuangxin v
                        ON l.项目编号 = v.编号
                    WHERE l.操作类型 = '有异议，上会'
                      AND ISNUMERIC(ISNULL(v.最高分, 0)) = 1
                      AND CAST(ISNULL(v.最高分, 0) AS INT) = 10
                )""")

        # 发起日期筛选
        if start_date_filter:
            where_conditions.append(f"发起时间 >= '{start_date_filter}'")

        if end_date_filter:
            where_conditions.append(f"发起时间 <= '{end_date_filter} 23:59:59'")

        if search_query:
            search_escaped = search_query.replace("'", "''")
            # 支持按编号搜索：如果输入中包含数字，则与编号进行模糊匹配
            number_cond = ""
            if any(ch.isdigit() for ch in search_query):
                # 去掉#等非数字字符，只保留数字部分
                digits_only = "".join(ch for ch in search_query if ch.isdigit())
                if digits_only:
                    number_cond = (
                        f" OR CHARINDEX('{digits_only}', ISNULL(CAST(编号 AS NVARCHAR(50)), '')) > 0"
                    )
            where_conditions.append(
                f"(CHARINDEX('{search_escaped}', ISNULL(CAST(标题 AS NVARCHAR(MAX)), '')) > 0 "
                f"OR CHARINDEX('{search_escaped}', ISNULL(CAST(发起人 AS NVARCHAR(200)), '')) > 0"
                f"{number_cond})"
            )

        if max_score_filters:
            allowed_scores = {0, 10, 20, 30, 50}
            parts = []
            for v in max_score_filters:
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                for p in s.replace('，', ',').split(','):
                    p = (p or '').strip()
                    if not p:
                        continue
                    if not p.isdigit():
                        continue
                    n = int(p)
                    if n in allowed_scores:
                        parts.append(n)
            uniq_scores = []
            seen = set()
            for n in parts:
                if n in seen:
                    continue
                seen.add(n)
                uniq_scores.append(n)
            if uniq_scores:
                in_clause = ", ".join(str(n) for n in uniq_scores)
                where_conditions.append(
                    "编号 IN ("
                    "SELECT DISTINCT [编号] "
                    "FROM v_quanyuanchuangxin "
                    f"WHERE (CASE WHEN ISNUMERIC(ISNULL([最高分], 0)) = 1 THEN CAST(ISNULL([最高分], 0) AS INT) ELSE 0 END) IN ({in_clause})"
                    ")"
                )

        if favorite_only:
            if not current_user_name:
                return jsonify({'success': False, 'message': '请先登录后再查看我的收藏'}), 401
            user_name_escaped = current_user_name.replace("'", "''")
            where_conditions.append(
                "编号 IN ("
                "SELECT DISTINCT TRY_CAST(bianhao AS INT) "
                "FROM chuangxin_shoucang "
                f"WHERE xingming = '{user_name_escaped}'"
                ")"
            )
        if favorite_category:
            category_escaped = favorite_category.replace("'", "''")
            where_conditions.append(
                "编号 IN ("
                "SELECT DISTINCT TRY_CAST(bianhao AS INT) "
                "FROM chuangxin_shoucang "
                f"WHERE ISNULL(fenlei, '') = N'{category_escaped}'"
                ")"
            )

        where_clause = ""
        if where_conditions:
            where_clause = "WHERE " + " AND ".join(where_conditions)

        # 1. 先获取总数（用于分页）
        count_sql = f"SELECT COUNT(*) FROM chuangxin_tibao1 {where_clause}"
        count_result = sf_db(count_sql)
        total_count = count_result[0] if count_result else 0

        columns_sql = """
            SELECT COLUMN_NAME
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME='chuangxin_tibao1'
              AND COLUMN_NAME IN (N'状态', N'图片详情', N'视频详情', N'创新类别', N'解决方案')
        """
        columns_rows = sf_db(columns_sql) or []
        existing_columns = set()
        for r in columns_rows:
            try:
                existing_columns.add((r[0] if isinstance(r, (list, tuple)) else r) or '')
            except Exception:
                pass

        status_expr = "状态" if "状态" in existing_columns else "''"
        image_expr = "图片详情" if "图片详情" in existing_columns else "''"
        video_expr = "视频详情" if "视频详情" in existing_columns else "''"
        category_expr = "创新类别" if "创新类别" in existing_columns else "''"
        solution_expr = "解决方案" if "解决方案" in existing_columns else "''"

        project_sql = f"""
            SELECT 编号,
                   标题,
                   内容,
                   发起人,
                   发起时间,
                   截止时间,
                   得分,
                   {status_expr} AS 状态,
                   {image_expr} AS 图片详情,
                   {video_expr} AS 视频详情,
                   {category_expr} AS 创新类别,
                   {solution_expr} AS 解决方案
            FROM chuangxin_tibao1
            {where_clause}
            ORDER BY 发起时间 DESC
            OFFSET {offset} ROWS FETCH NEXT {per_page} ROWS ONLY
        """
        projects_data = sf_db(project_sql)

        if not projects_data:
            _safe_print("⚠️ 没有查询到任何项目数据")
            return jsonify({
                'success': True,
                'data': [],
                'message': '暂无创新项目数据'
            })

        innovations = []
        _safe_print(f"🔄 开始处理 {len(projects_data)} 个项目...")

        favorite_set = set()
        if current_user_name and projects_data:
            try:
                bianhao_list = []
                for p in projects_data:
                    pid = p[0]
                    if pid is None:
                        continue
                    pid_text = str(pid).replace("'", "''")
                    bianhao_list.append(f"'{pid_text}'")
                if bianhao_list:
                    user_name_escaped = current_user_name.replace("'", "''")
                    fav_sql = f"""
                        SELECT DISTINCT bianhao
                        FROM chuangxin_shoucang
                        WHERE xingming = '{user_name_escaped}'
                          AND bianhao IN ({",".join(bianhao_list)})
                    """
                    fav_rows = sf_db(fav_sql) or []
                    for row in fav_rows:
                        if isinstance(row, (list, tuple)):
                            favorite_set.add(str(row[0]).strip())
                        else:
                            favorite_set.add(str(row).strip())
            except Exception as fav_err:
                _safe_print(f"⚠️ 查询收藏状态失败: {fav_err}")

        # 相似提案Top1（阈值80%）
        similar_top1_map = {}
        try:
            if projects_data:
                bianhao_list = []
                for p in projects_data:
                    pid = p[0]
                    if pid is None:
                        continue
                    pid_text = str(pid).replace("'", "''")
                    bianhao_list.append(f"'{pid_text}'")
                if bianhao_list:
                    sim_sql = f"""
                        SELECT x.bianhao, x.xiangsi1_bianhao, x.xiangsi1_fenshu, t.标题
                        FROM chuangxin_xiangliang x
                        LEFT JOIN chuangxin_tibao1 t
                               ON CAST(t.编号 AS NVARCHAR(100)) = x.xiangsi1_bianhao
                        WHERE x.bianhao IN ({",".join(bianhao_list)})
                          AND ISNULL(x.xiangsi1_fenshu, 0) >= 0.8
                    """
                    sim_rows = sf_db(sim_sql) or []
                    for row in sim_rows:
                        b = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
                        sb = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                        sc = float(row[2]) if len(row) > 2 and row[2] is not None else 0.0
                        st = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
                        if b and sb and sc >= 0.8:
                            similar_top1_map[b] = {
                                'bianhao': sb,
                                'title': st,
                                'score': round(sc, 4),
                                'score_percent': round(sc * 100, 2)
                            }
        except Exception as sim_err:
            _safe_print(f"⚠️ 查询相似提案失败: {sim_err}")

        for i, project in enumerate(projects_data):
            project_id = project[0]

            # ... 现有的流转记录查询代码保持不变 ...
            # 查询该项目的所有流转记录（按部门组长）
            try:
                flow_details_sql = f"""
                    SELECT 流转ID, 承接人, 状态, 分数, 处理备注, 流转时间, 处理时间
                    FROM chuangxin_liuzhuan1
                    WHERE 项目编号 = {project_id}
                    ORDER BY 流转次数 DESC, 流转时间 DESC
                """
                flow_details = sf_db(flow_details_sql)

                # 构建部门状态映射
                department_status = {}
                assignees_list = []

                for flow in flow_details:
                    flow_id, dept, status, score, notes, flow_time, handle_time = flow
                    if dept not in assignees_list:
                        assignees_list.append(dept)

                    if dept not in department_status:
                        department_status[dept] = {
                            'flow_id': flow_id,
                            'status': status or '待承接',
                            'score': score,
                            'notes': notes,
                            'handle_time': str(handle_time) if handle_time else ''
                        }

                # 计算整体状态
                status_priority = {'进行中': 4, '已完成': 3, '已拒绝': 2, '待承接': 1}
                overall_status = '待承接'
                current_handler = ''
                handler_notes = ''
                handle_time = ''

                for dept, info in department_status.items():
                    if status_priority.get(info['status'], 0) > status_priority.get(overall_status, 0):
                        overall_status = info['status']
                        current_handler = dept
                        handler_notes = info['notes'] or ''
                        handle_time = info['handle_time']


            except Exception as assignee_error:

                assignees_list = []
                department_status = {}
                overall_status = '待承接'
                current_handler = ''
                handler_notes = ''
                handle_time = ''

            def process_image_field(image_data):
                """处理图片字段数据，支持多个图片路径（分号分隔）"""
                if not image_data:
                    return None

                def _basename_anysep(p):
                    s = str(p or "").strip().strip('"\'')
                    if not s:
                        return ""
                    s = s.replace("\\", "/")
                    return s.split("/")[-1].strip()

                if isinstance(image_data, bytes):
                    # 如果是bytes类型，尝试解码为字符串
                    try:
                        decoded_path = image_data.decode('utf-8')
                        # 处理多个路径（分号分隔）
                        if ';' in decoded_path:
                            paths = decoded_path.split(';')
                            processed_paths = []
                            for path in paths:
                                path = path.strip()
                                if path:
                                    # 如果是完整路径，提取文件名
                                    if '\\' in path or '/' in path:
                                        processed_paths.append(_basename_anysep(path))
                                    else:
                                        processed_paths.append(path)
                            return ';'.join(processed_paths)
                        else:
                            # 单个路径处理
                            if '\\' in decoded_path or '/' in decoded_path:
                                return _basename_anysep(decoded_path)
                            return decoded_path
                    except UnicodeDecodeError:
                        _safe_print(f"⚠️ 图片数据无法解码为字符串，跳过显示")
                        return None
                elif isinstance(image_data, str):
                    if image_data.lower().startswith('0x'):
                        try:
                            image_data = bytes.fromhex(image_data[2:]).decode('utf-8')
                        except (ValueError, UnicodeDecodeError):
                            pass
                    # 处理多个路径（分号分隔）
                    if ';' in image_data:
                        paths = image_data.split(';')
                        processed_paths = []
                        for path in paths:
                            path = path.strip()
                            if path:
                                # 如果是完整路径，提取文件名
                                if '\\' in path or '/' in path:
                                    processed_paths.append(_basename_anysep(path))
                                else:
                                    processed_paths.append(path)
                        return ';'.join(processed_paths)
                    else:
                        # 单个路径处理
                        if '\\' in image_data or '/' in image_data:
                            return _basename_anysep(image_data)
                        return image_data
                else:
                    return str(image_data)

            def process_text_field(text_data):
                """处理文本字段数据，确保返回字符串类型"""
                if not text_data:
                    return ''

                if isinstance(text_data, bytes):
                    # 如果是bytes类型，尝试解码为字符串
                    try:
                        return text_data.decode('utf-8')
                    except UnicodeDecodeError:
                        _safe_print(f"⚠️ 文本数据无法解码为字符串，返回空字符串")
                        return ''
                elif isinstance(text_data, str):
                    return text_data
                else:
                    return str(text_data)

            # 在 process_score_field 函数中添加调试信息
            def process_score_field(score_data):
                """处理得分字段数据，确保返回数字类型"""

                if not score_data:
                    return 0

                if isinstance(score_data, bytes):
                    # 如果是bytes类型，尝试解码为字符串再转换为数字
                    try:
                        decoded_score = score_data.decode('utf-8')

                        result = float(decoded_score) if decoded_score else 0

                        return result
                    except (UnicodeDecodeError, ValueError) as e:
                        _safe_print(f"⚠️ 得分数据无法解码或转换为数字: {e}，返回0")
                        return 0
                elif isinstance(score_data, (int, float)):

                    return score_data
                elif isinstance(score_data, str):
                    try:
                        result = float(score_data) if score_data else 0

                        return result
                    except ValueError as e:
                        _safe_print(f"⚠️ 得分字符串无法转换为数字: {score_data}，错误: {e}，返回0")
                        return 0
                else:
                    try:
                        result = float(score_data)

                        return result
                    except (ValueError, TypeError) as e:
                        _safe_print(f"⚠️ 得分数据类型无法处理: {type(score_data)}，错误: {e}，返回0")
                        return 0

            # 在构建 innovation_data 之前添加调试信息

            # 查询委员会打分和操作类型
            committee_score = ''
            operation_type = ''
            committee_notes = ''
            try:
                committee_sql = f"""
                    SELECT TOP 1 委员会打分, 操作类型, 处理备注
                    FROM chuangxin_liuzhuan1
                    WHERE 项目编号 = {project_id}
                    AND (
                        (委员会打分 IS NOT NULL AND 委员会打分 != '' AND 委员会打分 != '-1')
                        OR (操作类型 IS NOT NULL AND 操作类型 != '')
                    )
                    ORDER BY ISNULL(处理时间, 流转时间) DESC, 流转ID DESC
                """
                committee_result = sf_db(committee_sql)
                if committee_result:
                    committee_score = str(committee_result[0][0]).strip() if committee_result[0][0] is not None else ''
                    operation_type = committee_result[0][1] if committee_result[0][1] else ''
                    remarks_raw = committee_result[0][2] if len(committee_result[0]) > 2 else ''
                    if remarks_raw and committee_score in {'', '-1', '0'}:
                        score_matches = re.findall(r'委员会打分[：:]\s*([0-9]+)', str(remarks_raw))
                        if score_matches:
                            committee_score = str(score_matches[-1]).strip()
                    if remarks_raw:
                        try:
                            remarks_text = str(remarks_raw)
                            idx_label = remarks_text.rfind('委员会备注：')
                            if idx_label != -1:
                                start_idx = remarks_text.rfind('[', 0, idx_label)
                                if start_idx == -1:
                                    start_idx = idx_label
                                segment = remarks_text[start_idx:]
                                end_idx = len(segment)
                                for sep in ['；', ';', '\n']:
                                    si = segment.find(sep)
                                    if si != -1:
                                        end_idx = si + 1
                                        break
                                committee_notes = segment[:end_idx].strip()
                        except Exception as parse_err:
                            _safe_print(f"⚠️ 解析委员会备注失败: {parse_err}")
            except Exception as e:
                _safe_print(f"⚠️ 查询委员会打分和操作类型失败: {e}")

            # 在 innovation_data 字典构建部分
            innovation_data = {
                'id': project[0],
                'title': process_text_field(project[1]),
                'content': process_text_field(project[2]),
                'initiator': process_text_field(project[3]),
                'initiation_time': str(project[4]) if project[4] else '',
                'deadline': str(project[5]) if project[5] else '',
                'score': process_score_field(project[6]),  # 🔧 修改这里，使用新的处理函数
                'status': overall_status,
                'department': ', '.join(assignees_list),
                'assignees': assignees_list,
                'handler': current_handler,
                'handle_time': handle_time,
                'handler_notes': handler_notes,
                'flow_count': len(assignees_list),
                'department_status': department_status,
                'image_path': process_image_field(project[8]) if len(project) > 8 else None,
                'video_path': process_image_field(project[9]) if len(project) > 9 else None,
                'project_type': process_text_field(project[10]) if len(project) > 10 else '未分类',
                'solution': process_text_field(project[11]) if len(project) > 11 else '',
                'committee_score': committee_score,  # 新增：委员会打分
                'operation_type': operation_type,  # 新增：操作类型
                'committee_notes': committee_notes,  # 新增：委员会备注（最新一条）
                'is_favorited': str(project[0]).strip() in favorite_set,
                'similar_top1': similar_top1_map.get(str(project[0]).strip())
            }

            # 直接添加到结果列表（筛选已在SQL中完成）
            innovations.append(innovation_data)

        # 计算总页数
        total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 0

        return jsonify({
            'success': True,
            'data': innovations,
            'total': total_count,
            'page': page,
            'per_page': per_page,
            'total_pages': total_pages
        })

    except Exception as e:
        _safe_print(f"💥 获取创新项目列表失败: {e}")
        _safe_print(f"📍 错误类型: {type(e).__name__}")
        import traceback
        _safe_print(f"📋 错误堆栈: {traceback.format_exc()}")

        return jsonify({
            'success': False,
            'message': f'获取创新项目列表失败: {str(e)}',
            'error_type': type(e).__name__
        }), 500


@app.route('/api/toggle_favorite', methods=['POST'])
def toggle_favorite():
    """收藏/取消收藏创新提案"""
    try:
        current_user_name = _get_current_feishu_user_name()
        if not current_user_name:
            return jsonify({'success': False, 'message': '请先登录'}), 401

        data = request.get_json(silent=True) or {}
        innovation_id_raw = str(data.get('innovation_id') or '').strip()
        if not innovation_id_raw or not innovation_id_raw.isdigit():
            return jsonify({'success': False, 'message': '提案编号无效'}), 400
        innovation_id = innovation_id_raw

        favorite_raw = data.get('favorite')
        desired_state = None
        if isinstance(favorite_raw, bool):
            desired_state = favorite_raw
        elif favorite_raw is not None:
            desired_state = str(favorite_raw).strip().lower() in {'1', 'true', 'yes'}

        user_name_escaped = current_user_name.replace("'", "''")
        innovation_id_escaped = innovation_id.replace("'", "''")

        exists_sql = f"""
            SELECT TOP 1 1
            FROM chuangxin_shoucang
            WHERE xingming = '{user_name_escaped}'
              AND bianhao = '{innovation_id_escaped}'
              AND ISNULL(fenlei, N'收藏') IN (N'', N'收藏')
        """
        exists_rows = sf_db(exists_sql) or []
        is_favorited = bool(exists_rows)

        target_state = (not is_favorited) if desired_state is None else desired_state
        if target_state and (not is_favorited):
            now_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            insert_sql = f"""
                INSERT INTO chuangxin_shoucang (xingming, bianhao, riqi, fenlei)
                VALUES ('{user_name_escaped}', '{innovation_id_escaped}', '{now_text}', N'收藏')
            """
            dui_db(insert_sql)
            is_favorited = True
        elif (not target_state) and is_favorited:
            delete_sql = f"""
                DELETE FROM chuangxin_shoucang
                WHERE xingming = '{user_name_escaped}'
                  AND bianhao = '{innovation_id_escaped}'
                  AND ISNULL(fenlei, N'收藏') IN (N'', N'收藏')
            """
            dui_db(delete_sql)
            is_favorited = False

        return jsonify({
            'success': True,
            'is_favorited': is_favorited,
            'message': '收藏成功' if is_favorited else '已取消收藏'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'收藏操作失败: {str(e)}'}), 500


@app.route('/api/add_favorite_category', methods=['POST'])
def add_favorite_category():
    """给创新提案添加分类收藏记录"""
    try:
        current_user_name = _get_current_feishu_user_name()
        if not current_user_name:
            return jsonify({'success': False, 'message': '请先登录'}), 401

        data = request.get_json(silent=True) or {}
        innovation_id_raw = str(data.get('innovation_id') or '').strip()
        fenlei = str(data.get('fenlei') or '').strip()
        allowed_categories = {'长期看效果', '暂未有解决方案'}
        if not innovation_id_raw or not innovation_id_raw.isdigit():
            return jsonify({'success': False, 'message': '提案编号无效'}), 400
        if fenlei not in allowed_categories:
            return jsonify({'success': False, 'message': '分类无效'}), 400

        user_name_escaped = current_user_name.replace("'", "''")
        innovation_id_escaped = innovation_id_raw.replace("'", "''")
        fenlei_escaped = fenlei.replace("'", "''")
        now_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        exists_sql = f"""
            SELECT TOP 1 1
            FROM chuangxin_shoucang
            WHERE xingming = '{user_name_escaped}'
              AND bianhao = '{innovation_id_escaped}'
              AND ISNULL(fenlei, '') = N'{fenlei_escaped}'
        """
        exists_rows = sf_db(exists_sql) or []
        if exists_rows:
            dui_db(f"""
                UPDATE chuangxin_shoucang
                SET riqi = '{now_text}'
                WHERE xingming = '{user_name_escaped}'
                  AND bianhao = '{innovation_id_escaped}'
                  AND ISNULL(fenlei, '') = N'{fenlei_escaped}'
            """)
        else:
            dui_db(f"""
                INSERT INTO chuangxin_shoucang (xingming, bianhao, riqi, fenlei)
                VALUES ('{user_name_escaped}', '{innovation_id_escaped}', '{now_text}', N'{fenlei_escaped}')
            """)

        return jsonify({
            'success': True,
            'message': f'已加入“{fenlei}”'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'分类记录失败: {str(e)}'}), 500


@app.route('/api/transfer_innovation', methods=['POST'])
def transfer_innovation():
    """流转创新项目到其他部门"""
    try:
        data = request.get_json()
        project_id = data['project_id']
        new_department = data['new_department']
        current_handler = data['current_handler']
        transfer_notes = data.get('transfer_notes', '')

        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 1. 获取当前项目的最大流转次数
        max_flow_sql = f"""
            SELECT MAX(流转次数) FROM chuangxin_liuzhuan1 
            WHERE 项目编号 = {project_id}
        """
        max_flow_result = sf_db(max_flow_sql)
        next_flow_number = (max_flow_result[0] if max_flow_result and max_flow_result[0] else 0) + 1

        # 2. 获取项目发起人
        project_sql = f"SELECT 发起人 FROM chuangxin_tibao1 WHERE 编号 = {project_id}"
        project_info = sf_db(project_sql)
        if not project_info:
            raise Exception("项目不存在")

        initiator = project_info[0]

        # 3. 插入新的流转记录
        new_department_escaped = new_department.replace("'", "''")
        initiator_escaped = initiator.replace("'", "''")

        flow_sql = f"""
            INSERT INTO chuangxin_liuzhuan1 (项目编号, 发起人, 承接人, 流转次数, 流转时间, 状态, 流转备注)
            VALUES ({project_id}, '{initiator_escaped}', '{new_department_escaped}', {next_flow_number}, '{current_time}', '待承接', '{transfer_notes}')
        """

        dui_db(flow_sql)

        # 4. 更新当前流转记录状态为已流转
        update_current_sql = f"""
            UPDATE chuangxin_liuzhuan1 
            SET 状态 = '已流转', 处理时间 = '{current_time}', 处理备注 = '流转至{new_department}'
            WHERE 项目编号 = {project_id} AND 承接人 = '{current_handler}' AND 状态 = '待承接'
        """

        dui_db(update_current_sql)

        return jsonify({
            'success': True,
            'message': f'项目已成功流转至{new_department}，流转次数：{next_flow_number}'
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'流转失败: {str(e)}'
        }), 500


@app.route('/api/request_help', methods=['POST'])
def request_help():
    """请求帮助API - 承接人可以请求其他人帮助"""
    try:
        data = request.get_json()
        project_id = data['project_id']
        current_user = data['current_user']  # 当前承接人
        help_users = data['help_users']  # 请求帮助的人员列表

        # 获取当前最大流转次数
        max_flow_sql = f"SELECT MAX(流转次数) FROM chuangxin_liuzhuan1 WHERE 项目编号 = {project_id}"
        max_flow_result = sf_db(max_flow_sql)
        next_flow_num = (max_flow_result[0][0] or 0) + 1

        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 为每个帮助人员创建流转记录
        for help_user in help_users:
            help_user = help_user.strip().replace("'", "''")
            current_user_escaped = current_user.replace("'", "''")

            flow_sql = f"""
                INSERT INTO chuangxin_liuzhuan1 (项目编号, 发起人, 承接人, 流转次数, 流转时间, 状态)
                VALUES ({project_id}, '{current_user_escaped}', '{help_user}', {next_flow_num}, '{current_time}', '待承接')
            """
            dui_db(flow_sql)

        return jsonify({
            'success': True,
            'message': '帮助请求发送成功！'
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'请求帮助失败: {str(e)}'
        }), 500


@app.route('/api/get_flow_details/<int:project_id>', methods=['GET'])
def get_flow_details(project_id):
    """获取项目流转详情"""
    try:
        sql = f"""
            SELECT 
                流转ID,
                发起人,
                承接人,
                流转次数,
                流转时间,
                状态,
                处理时间,
                处理备注
            FROM chuangxin_liuzhuan1 
            WHERE 项目编号 = {project_id}
            ORDER BY 流转次数, 流转时间
        """

        flow_data = sf_db(sql)
        flows = []

        for row in flow_data:
            flows.append({
                'flow_id': row[0],
                'initiator': row[1],
                'assignee': row[2],
                'flow_number': row[3],
                'flow_time': row[4],
                'status': row[5],
                'handle_time': row[6],
                'handle_notes': row[7]
            })

        return jsonify({
            'success': True,
            'data': flows
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'获取流转详情失败: {str(e)}'
        }), 500


@app.route('/api/get_project_departments/<int:project_id>', methods=['GET'])
def get_project_departments(project_id):
    """获取项目的承接部门列表"""
    try:
        # 查询该项目的所有流转记录
        sql = f"""
            SELECT 
                流转ID,
                承接人 as department,
                状态,
                分数,
                处理备注,
                流转时间,
                处理时间
            FROM chuangxin_liuzhuan1 
            WHERE 项目编号 = {project_id}
            ORDER BY 流转次数, 流转时间
        """

        flow_data = sf_db(sql)
        departments = []

        for row in flow_data:
            departments.append({
                'flow_id': row[0],
                'department': row[1],
                'status': row[2] or '待承接',
                'score': row[3],
                'notes': row[4],
                'flow_time': str(row[5]) if row[5] else '',
                'handle_time': str(row[6]) if row[6] else ''
            })

        return jsonify({
            'success': True,
            'data': departments
        })

    except Exception as e:
        _safe_print(f"❌ 获取项目部门列表失败: {e}")
        return jsonify({
            'success': False,
            'message': f'获取部门列表失败: {str(e)}'
        }), 500


@app.route('/api/handle_innovation', methods=['POST'])
def handle_innovation():
    """处理创新项目"""
    try:
        data = request.get_json()
        innovation_id = data['innovation_id']
        flow_id = data['flow_id']
        status = data['status']
        handler = data['handler']
        notes = data.get('notes', '')
        score = data.get('score', '')
        committee_score = data.get('committee_score', '')  # 新增：委员会打分
        operation_type = data.get('operation_type', '处理')

        _safe_print(f"🔄 开始处理创新项目")
        _safe_print(f"📋 参数: innovation_id={innovation_id}, flow_id={flow_id}, status={status}")
        _safe_print(f"👤 处理人: {handler}, 备注: {notes}, 部门评分: {score}")

        # 参数验证
        if not all([innovation_id, flow_id, status, handler]):
            return jsonify({
                'success': False,
                'message': '缺少必要参数'
            }), 400

        # 获取项目信息
        project_sql = f"SELECT 标题, 发起人 FROM chuangxin_tibao1 WHERE 编号 = {innovation_id}"
        _safe_print(f"🔍 查询项目信息SQL: {project_sql}")
        project_info = sf_db(project_sql)

        if not project_info:
            return jsonify({
                'success': False,
                'message': '项目不存在'
            }), 404

        title, initiator = project_info[0]
        _safe_print(f"📝 项目信息: 标题={title}, 发起人={initiator}")

        # 处理评分等级转换
        numeric_score = 0
        score_to_store = score  # 默认存储原始评分
        final_score_for_display = score  # 用于通知显示的评分
        committee_score_to_store = ''

        # 评分映射
        score_mapping = {
            'A+': 100,
            'A': 50,
            'B': 30,
            'C': 20,
            'D': 10,
            'E': 0,
        }

        # 优先处理委员会打分，如果有委员会打分则覆盖部门打分
        if committee_score:
            if committee_score in score_mapping:
                numeric_score = score_mapping[committee_score]
                score_to_store = str(numeric_score)  # 存储委员会评分的数字分数
                committee_score_to_store = committee_score  # 存储委员会评分等级
                final_score_for_display = committee_score  # 通知中显示委员会评分
                _safe_print(f"📊 委员会评分转换: {committee_score} -> {numeric_score}分（覆盖部门评分）")
            elif committee_score.isdigit():
                numeric_score = int(committee_score)
                score_to_store = committee_score
                committee_score_to_store = committee_score
                final_score_for_display = committee_score
                _safe_print(f"📊 委员会数字评分: {numeric_score}分（覆盖部门评分）")
        elif score:
            # 没有委员会打分时，使用部门打分
            if score in score_mapping:
                numeric_score = score_mapping[score]
                score_to_store = str(numeric_score)  # 存储数字分数而不是字母等级
                final_score_for_display = score
                _safe_print(f"📊 部门评分转换: {score} -> {numeric_score}分")
            elif score.isdigit():
                numeric_score = int(score)
                score_to_store = score  # 已经是数字，直接存储
                final_score_for_display = score
                _safe_print(f"📊 部门数字评分: {numeric_score}分")

        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 转义特殊字符
        handler_escaped = handler.replace("'", "''")
        # 在handle_innovation函数中，在获取notes后添加
        notes = data.get('handler_notes', '').strip()

        # 如果notes为空，返回错误
        if not notes:
            _safe_print(f"❌ 处理备注为空，拒绝处理")
            return jsonify({
                'success': False,
                'message': '处理备注不能为空，请填写处理意见'
            }), 400

        score_escaped = score_to_store.replace("'", "''") if score_to_store else ''  # 使用转换后的数字分数

        # 更新流转记录 - 存储数字分数和委员会打分，使用CASE语句进行真正的追加
        committee_score_escaped = committee_score_to_store.replace("'", "''") if committee_score_to_store else ''

        # 构建新的处理备注内容
        if committee_score:
            # 有委员会打分时的备注格式
            committee_score_display = committee_score_to_store if committee_score_to_store else committee_score
            new_note_content = f"[{current_time}] {handler.replace("'", "''")}：{notes.replace("'", "''")}[{current_time}] 委员会打分：{committee_score_display}分； "
        else:
            # 没有委员会打分时的备注格式
            new_note_content = f"[{current_time}] {handler.replace("'", "''")}：{notes.replace("'", "''")}； "

        flow_update_sql = f"""
            UPDATE chuangxin_liuzhuan1
            SET 状态 = '{status}',
                处理备注 = CASE 
                    WHEN 处理备注 IS NULL OR 处理备注 = '' THEN '{new_note_content}'
                    ELSE 处理备注 + '{new_note_content}'
                END,
                分数 = '{score_escaped}',
                委员会打分 = CASE
                    WHEN '{committee_score_escaped}' = '' THEN 委员会打分
                    ELSE '{committee_score_escaped}'
                END,
                处理时间 = '{current_time}',
                操作类型='{operation_type.replace("'", "''")}'
            WHERE 流转ID = {flow_id} AND 项目编号 = {innovation_id}
        """
        _safe_print(f"🔄 更新流转记录SQL: {flow_update_sql}")

        try:
            dui_db(flow_update_sql)
            _safe_print(f"✅ 流转记录更新成功")
        except Exception as flow_error:
            _safe_print(f"❌ 流转记录更新失败: {flow_error}")
            return jsonify({
                'success': False,
                'message': f'流转记录更新失败: {str(flow_error)}'
            }), 500

        # 更新创新项目总得分（优先使用委员会打分，否则取所有部门评分的最高分）
        if numeric_score > 0:
            # 首先查询是否有委员会打分
            committee_score_sql = f"""
                SELECT TOP 1 CASE 
                    WHEN ISNUMERIC(分数) = 1 THEN CAST(分数 AS INT)
                    ELSE 0
                END as score FROM chuangxin_liuzhuan1 
                WHERE 项目编号 = {innovation_id} AND 委员会打分 IS NOT NULL AND 委员会打分 != '' AND 委员会打分 != '-1'
                ORDER BY 处理时间 DESC
            """
            committee_score_result = sf_db(committee_score_sql)

            final_committee_score = 0
            if committee_score_result:
                if isinstance(committee_score_result, list) and len(committee_score_result) > 0:
                    first_row = committee_score_result[0]
                    if isinstance(first_row, (list, tuple)) and len(first_row) > 0:
                        value = first_row[0]
                    else:
                        value = first_row
                    if value is not None:
                        try:
                            final_committee_score = int(value)
                        except (ValueError, TypeError):
                            final_committee_score = 0
                elif isinstance(committee_score_result, (int, float)):
                    final_committee_score = int(committee_score_result)

            if final_committee_score > 0:
                final_project_score = final_committee_score
                _safe_print(f"📊 使用委员会打分作为项目最终得分: {final_project_score}分")
            else:
                max_score_sql = f"""
                    SELECT MAX(CASE 
                        WHEN ISNUMERIC(分数) = 1 THEN CAST(分数 AS INT)
                        ELSE 0
                    END) FROM chuangxin_liuzhuan1 
                    WHERE 项目编号 = {innovation_id} AND 分数 IS NOT NULL AND 分数 != ''
                """
                max_score_result = sf_db(max_score_sql)

                # 修正数据访问逻辑
                if max_score_result:
                    if isinstance(max_score_result, list) and len(max_score_result) > 0:
                        # 如果返回的是列表格式 [[value]] 或 [(value,)]
                        if isinstance(max_score_result[0], (list, tuple)) and len(max_score_result[0]) > 0:
                            final_project_score = max_score_result[0][0] if max_score_result[0][0] is not None else 0
                        else:
                            final_project_score = max_score_result[0] if max_score_result[0] is not None else 0
                    elif isinstance(max_score_result, (int, float)):
                        # 如果直接返回数值
                        final_project_score = max_score_result
                    else:
                        final_project_score = 0
                else:
                    final_project_score = 0
                _safe_print(f"📊 使用部门评分最高分作为项目最终得分: {final_project_score}分")

            # 更新项目总得分
            update_project_sql = f"""
                UPDATE chuangxin_tibao1
                SET 得分 = {final_project_score}
                WHERE 编号 = {innovation_id}
            """
            _safe_print(f"🔄 更新项目得分SQL: {update_project_sql}")

            try:
                dui_db(update_project_sql)
                _safe_print(f"✅ 项目得分更新成功，最终得分: {final_project_score}")
            except Exception as sql_error:
                _safe_print(f"❌ 项目得分更新失败: {sql_error}")
                # 不影响主流程，继续执行

        # 发送飞书通知（异步处理，不影响主流程）
        try:
            _safe_print(f"📢 开始发送飞书通知")
            message_service = get_message_service()
            message_service.send_handle_notification(
                title=title,
                handler=handler,
                status=status,
                notes=notes,
                score=final_score_for_display,  # 通知中显示最终评分（委员会打分优先）
                submitter=initiator  # 添加提案提交者信息
            )
            _safe_print(f"✅ 飞书通知发送成功")
        except Exception as e:
            _safe_print(f"❌ 飞书通知发送失败: {e}")
            # 不影响主流程，继续执行

        _safe_print(f"🎉 处理完成，返回成功响应")
        return jsonify({
            'success': True,
            'message': '处理成功！'
        })

    except Exception as e:
        _safe_print(f"💥 处理过程中发生异常: {str(e)}")
        _safe_print(f"📍 异常类型: {type(e).__name__}")
        import traceback
        _safe_print(f"📋 完整错误堆栈:\n{traceback.format_exc()}")
        return jsonify({
            'success': False,
            'message': f'处理失败: {str(e)}'
        }), 500


@app.route('/api/update_committee_score', methods=['POST'])
def update_committee_score():
    """委员会打分直接更新接口"""
    try:
        data = request.get_json()
        innovation_id = data.get('innovation_id')
        committee_score = data.get('committee_score')
        committee_notes = (data.get('committee_notes') or '').strip()

        _safe_print(f"🔄 委员会打分直接更新: 项目ID={innovation_id}, 评分={committee_score}")

        # 参数验证
        if not innovation_id or not committee_score:
            return jsonify({
                'success': False,
                'message': '缺少必要参数'
            }), 400

        # 验证用户权限
        raw_name = session.get('feishu_user_name', '')
        name_parts = str(raw_name).split('（', 1) if raw_name else []
        feishu_user_name = name_parts[0].strip() if name_parts else str(raw_name).strip()
        committee_members = ['周俊成', '陶晓飞', '孙军', '毕景春', '李昌瀚', '蔡晶', '韩雅俊','孙洁','陈子烨']

        if feishu_user_name not in committee_members:
            return jsonify({
                'success': False,
                'message': '无权限进行委员会打分'
            }), 403

        # 评分映射
        score_mapping = {
            'A+': 100,
            'A': 50,
            'B': 30,
            'C': 20,
            'D': 10,
            'E': 0,
        }

        # 转换评分
        if committee_score in score_mapping:
            numeric_score = score_mapping[committee_score]
        elif committee_score.isdigit():
            numeric_score = int(committee_score)
        else:
            return jsonify({
                'success': False,
                'message': '无效的评分值'
            }), 400

        # 获取当前时间
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 构建委员会打分信息 - 显示数字评分而不是等级
        committee_info = f"[{current_time}] 委员会打分：{numeric_score}分;"
        committee_info_escaped = committee_info.replace("'", "''")

        # 构建委员会备注信息（可选）
        notes_info = ''
        if committee_notes:
            notes_info = f"[{current_time}] 委员会备注：{committee_notes};"
        notes_info_escaped = notes_info.replace("'", "''")

        # 空备注时的合并字符串
        new_remarks_when_empty = f"[{current_time}] 委员会打分：{numeric_score}分" + (
            notes_info if committee_notes else '')
        new_remarks_when_empty_escaped = new_remarks_when_empty.replace("'", "''")

        _safe_print(f"📝 追加委员会打分信息到已有备注: {committee_info}")

        # 使用SQL字符串拼接来真正追加内容而不是覆盖 - 只更新最新的流转记录
        update_sql = f"""
            UPDATE chuangxin_liuzhuan1 
            SET 委员会打分 = {numeric_score},
                处理备注 = CASE 
                    WHEN 处理备注 IS NULL OR 处理备注 = '' 
                    THEN '{new_remarks_when_empty_escaped}'
                    ELSE 处理备注 + '{committee_info_escaped}'{(' + \'' + notes_info_escaped + '\'') if committee_notes else ''}
                END
            WHERE 项目编号 = {innovation_id}
            AND 流转ID = (
                SELECT TOP 1 流转ID 
                FROM chuangxin_liuzhuan1 
                WHERE 项目编号 = {innovation_id}
                ORDER BY 处理时间 DESC, 流转ID DESC
            )
        """

        _safe_print(f"📝 更新SQL: {update_sql}")

        result = dui_db(update_sql)

        _safe_print(f"✅ 委员会打分更新成功: 项目{innovation_id} -> {committee_score}({numeric_score}分)")

        try:
            project_sql = f"SELECT 标题, 发起人 FROM chuangxin_tibao1 WHERE 编号 = {innovation_id}"
            _safe_print(f"🔍 查询项目信息SQL: {project_sql}")
            project_info = sf_db(project_sql)
            if project_info:
                title, initiator = project_info[0]
                message_service = get_message_service()
                score_text = f"{committee_score}({numeric_score}分)" if committee_score else f"{numeric_score}分"
                notes_for_notify = committee_notes
                handler_name = feishu_user_name
                message_service.send_handle_notification(
                    title=title,
                    handler=handler_name,
                    status="委员会打分更新",
                    notes=notes_for_notify,
                    score=score_text,
                    submitter=initiator
                )
                _safe_print("✅ 委员会打分通知已发送给发起人")
            else:
                _safe_print("⚠️ 未找到对应的创新项目，无法发送通知")
        except Exception as notify_error:
            _safe_print(f"❌ 委员会打分通知发送失败: {notify_error}")

        return jsonify({
            'success': True,
            'message': f'委员会打分更新成功: {committee_score}({numeric_score}分)'
        })

    except Exception as e:
        _safe_print(f"❌ 委员会打分更新失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'更新失败: {str(e)}'
        }), 500


@app.route('/api/submit_to_meeting', methods=['POST'])
def submit_to_meeting():
    """一键上会接口 - 将操作类型更新为'有异议，上会'"""
    try:
        data = request.get_json()
        innovation_id = str(data.get('innovation_id') or '').strip()

        _safe_print(f"🔄 一键上会: 项目ID={innovation_id}")

        # 参数验证
        if not innovation_id or not innovation_id.isdigit():
            return jsonify({
                'success': False,
                'message': '提案编号无效'
            }), 400

        # 验证用户权限 - 所有登录用户都可以使用此功能
        feishu_user_name = session.get('feishu_user_name', '')

        if not feishu_user_name:
            return jsonify({
                'success': False,
                'message': '请先登录'
            }), 401

        # 获取当前时间
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 更新操作类型为"有异议，上会"
        update_sql = f"""
            UPDATE chuangxin_liuzhuan1
            SET 操作类型 = '有异议，上会',
                处理时间 = '{current_time}'
            WHERE 项目编号 = {innovation_id}
            AND 流转ID = (
                SELECT TOP 1 流转ID 
                FROM chuangxin_liuzhuan1 
                WHERE 项目编号 = {innovation_id}
                ORDER BY 处理时间 DESC, 流转ID DESC
            )
        """

        _safe_print(f"📝 一键上会SQL: {update_sql}")

        result = dui_db(update_sql)

        _safe_print(f"✅ 一键上会成功: 项目{innovation_id} -> 有异议，上会")
        return jsonify({
            'success': True,
            'message': '已成功标记为"有异议，上会"'
        })

    except Exception as e:
        _safe_print(f"❌ 一键上会失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'操作失败: {str(e)}'
        }), 500


@app.route('/api/update_user_points_summary', methods=['POST'])
def update_user_points_summary():
    """统计并更新用户积分汇总到创新_消费表"""
    try:
        data = request.get_json()
        user_name = data.get('user_name')
        current_year = datetime.now().year

        if not user_name:
            return jsonify({
                'success': False,
                'message': '用户名不能为空'
            }), 400

        # 1. 计算年度积分（当年获得的总积分）
        annual_points_sql = f"""
            SELECT SUM(最高分) as annual_points
            FROM v_quanyuanchuangxin 
            WHERE 发起人 = '{user_name}' 
            AND YEAR(发起时间) = {current_year}
            AND 最高分 IS NOT NULL AND 最高分 > 0
        """
        annual_result = sf_db(annual_points_sql)
        annual_points = annual_result[0][0] if annual_result else 0

        # 2. 计算历史总积分（所有年份获得的总积分）
        total_points_sql = f"""
            SELECT COALESCE(SUM(最高分), 0) as total_points
            FROM v_quanyuanchuangxin 
            WHERE 发起人 = '{user_name}'
            AND 最高分 IS NOT NULL AND 最高分 > 0
        """
        total_result = sf_db(total_points_sql)
        total_earned_points = total_result[0][0] if total_result else 0

        # 3. 计算已花费积分
        spent_points_sql = f"""
            SELECT COALESCE(SUM(花费积分), 0) as spent_points
            FROM 创新_消费 
            WHERE 使用人 = '{user_name}'
        """
        spent_result = sf_db(spent_points_sql)
        spent_points = spent_result[0][0] if spent_result else 0

        # 4. 计算当前积分（历史总积分 - 花费积分）
        current_points = total_earned_points - spent_points

        # 5. 检查用户是否已存在记录
        check_user_sql = f"""
            SELECT COUNT(*) FROM 创新_消费 
            WHERE 使用人 = '{user_name}' AND 备注 = '积分汇总'
        """
        user_exists = (sf_db(check_user_sql, single=True) or 0) > 0

        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if user_exists:
            # 更新现有记录
            update_sql = f"""
                UPDATE 创新_消费 
                SET 年度积分 = {annual_points},
                    历史总积分 = {total_earned_points},
                    当前积分 = {current_points},
                    日期 = '{current_time}'
                WHERE 使用人 = '{user_name}' AND 备注 = '积分汇总'
            """
            dui_db(update_sql)
        else:
            # 插入新记录
            insert_sql = f"""
                INSERT INTO 创新_消费 (使用人, 日期, 花费积分, 备注, 年度积分, 历史总积分, 当前积分)
                VALUES ('{user_name}', '{current_time}', 0, '积分汇总', {annual_points}, {total_earned_points}, {current_points})
            """
            dui_db(insert_sql)

        return jsonify({
            'success': True,
            'message': '积分统计更新成功',
            'data': {
                'user_name': user_name,
                'annual_points': annual_points,
                'total_earned_points': total_earned_points,
                'spent_points': spent_points,
                'current_points': current_points
            }
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'积分统计更新失败: {str(e)}'
        }), 500


@app.route('/api/batch_update_all_users_points', methods=['POST'])
def batch_update_all_users_points():
    """批量更新所有用户的积分统计"""
    try:
        # 获取所有有积分记录的用户
        users_sql = """
                    SELECT DISTINCT 发起人
                    FROM chuangxin_liuzhuan1
                    WHERE 发起人 IS NOT NULL \
                      AND 发起人 != ''
                    UNION
                    SELECT DISTINCT 使用人
                    FROM 创新_消费
                    WHERE 使用人 IS NOT NULL \
                      AND 使用人 != '' \
                    """
        users_result = sf_db(users_sql)

        if not users_result:
            return jsonify({
                'success': True,
                'message': '没有找到需要更新的用户',
                'updated_count': 0
            })

        updated_count = 0
        current_year = datetime.now().year
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        for user_row in users_result:
            user_name = user_row[0]

            # 计算该用户的积分统计
            annual_points_sql = f"""
                SELECT COALESCE(SUM(最高分), 0) as annual_points
                FROM v_quanyuanchuangxin
                WHERE 发起人 = '{user_name}' 
                AND YEAR(发起时间) = {current_year}
                AND 最高分 IS NOT NULL AND 最高分 > 0
            """
            annual_result = sf_db(annual_points_sql)
            annual_points = annual_result[0][0] if annual_result else 0

            total_points_sql = f"""
                SELECT COALESCE(SUM(最高分), 0) as total_points
                FROM v_quanyuanchuangxin 
                WHERE 发起人 = '{user_name}'
                AND 最高分 IS NOT NULL AND 最高分 > 0
            """
            total_result = sf_db(total_points_sql)
            total_earned_points = total_result[0][0] if total_result else 0

            spent_points_sql = f"""
                SELECT COALESCE(SUM(花费积分), 0) as spent_points
                FROM 创新_消费 
                WHERE 使用人 = '{user_name}'
            """
            spent_result = sf_db(spent_points_sql)
            spent_points = spent_result[0][0] if spent_result else 0

            current_points = total_earned_points - spent_points

            # 检查是否已存在汇总记录
            check_user_sql = f"""
                SELECT COUNT(*) FROM 创新_消费 
                WHERE 使用人 = '{user_name}' AND 备注 = '积分汇总'
            """
            user_exists = (sf_db(check_user_sql, single=True) or 0) > 0

            if user_exists:
                update_sql = f"""
                    UPDATE 创新_消费 
                    SET 年度积分 = {annual_points},
                        历史总积分 = {total_earned_points},
                        当前积分 = {current_points},
                        日期 = '{current_time}'
                    WHERE 使用人 = '{user_name}' AND 备注 = '积分汇总'
                """
                dui_db(update_sql)
            else:
                insert_sql = f"""
                    INSERT INTO 创新_消费 (使用人, 日期, 花费积分, 备注, 年度积分, 历史总积分, 当前积分)
                    VALUES ('{user_name}', '{current_time}', 0, '积分汇总', {annual_points}, {total_earned_points}, {current_points})
                """
                dui_db(insert_sql)

            updated_count += 1

        return jsonify({
            'success': True,
            'message': f'批量更新完成，共更新 {updated_count} 个用户的积分统计',
            'updated_count': updated_count
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'批量更新失败: {str(e)}'
        }), 500


# 积分系统API
@app.route('/api/get_user_points', methods=['GET'])
def get_user_points():
    """获取用户积分API - 修复查询结果处理"""
    try:
        user_name = (request.args.get('user_name') or '').strip()
        if not user_name:
            user_name = _get_current_feishu_user_name()

        if not user_name:
            _safe_print("[ERROR] 未登录或缺少用户名参数")
            return jsonify({'success': False, 'message': '请先登录飞书后再查看积分'}), 401
        user_esc = user_name.replace("'", "''")

        current_year = datetime.now().year

        # 统一的结果处理函数
        def extract_value(result):
            """从查询结果中提取数值"""
            if not result:
                return 0
            if isinstance(result, list):
                if len(result) > 0:
                    if isinstance(result[0], (list, tuple)):
                        return result[0][0] if len(result[0]) > 0 else 0
                    else:
                        return result[0]
                return 0
            elif isinstance(result, (int, float)):
                return result
            else:
                return 0

        # 1. 计算年度积分 - 放宽条件
        annual_points_sql = f"""
              SELECT COALESCE(SUM(最高分), 0) as annual_points
            FROM v_quanyuanchuangxin 
            WHERE 发起人 = '{user_esc}'
            AND YEAR(发起时间) = {current_year}
            AND 最高分 IS NOT NULL AND 最高分 > 0
        """

        try:
            annual_result = sf_db(annual_points_sql)

            annual_points = extract_value(annual_result)

        except Exception as e:
            _safe_print(f"[ERROR] 年度积分查询失败: {e}")
            annual_points = 0

        # 2. 计算历史总积分 - 放宽条件
        total_points_sql = f"""
            SELECT COALESCE(SUM(最高分), 0) as annual_points
            FROM v_quanyuanchuangxin 
            WHERE 发起人 = '{user_esc}'

            AND 最高分 IS NOT NULL AND 最高分 > 0
        """

        try:
            total_result = sf_db(total_points_sql)

            total_earned_points = extract_value(total_result)

        except Exception as e:
            _safe_print(f"[ERROR] 历史总积分查询失败: {e}")
            total_earned_points = 0

        # 3. 计算已花费积分
        spent_points_sql = f"""
            SELECT COALESCE(SUM(花费积分), 0) as spent_points
            FROM 创新_消费 
            WHERE 使用人 = '{user_esc}'
        """

        try:
            spent_result = sf_db(spent_points_sql)

            spent_points = extract_value(spent_result)

        except Exception as e:
            _safe_print(f"[ERROR] 已花费积分查询失败: {e}")
            spent_points = 0

        # 4. 计算当前积分
        current_points = total_earned_points - spent_points

        points_data = {
            'current_points': current_points,
            'annual_points': annual_points,
            'total_earned_points': total_earned_points
        }

        return jsonify({
            'success': True,
            'data': points_data
        })

    except Exception as e:
        error_msg = f'获取用户积分失败: {str(e)}'
        _safe_print(f"[ERROR] {error_msg}")
        _safe_print(f"[ERROR] 异常详情: {type(e).__name__}: {e}")
        import traceback
        _safe_print(f"[ERROR] 堆栈跟踪: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'message': error_msg
        }), 500


@app.route('/api/get_rewards', methods=['GET'])
def get_rewards():
    """获取奖品列表API"""
    try:

        category = request.args.get('category')
        manage_mode = request.args.get('manage_mode', 'false').lower() == 'true'

        # 根据实际表结构修改SQL查询
        sql = """
              SELECT id,
                     奖品                 as name,
                     ''                   as description,
                     奖品对应积分         as points_required,
                     库存                 as stock,
                     ISNULL(奖励图片, '') as image_path,
                     '默认'               as category,
                     1                    as is_active
              FROM 创新_奖品表
              ORDER BY 奖品对应积分 ASC
              """

        rewards_data = sf_db(sql)

        rewards = []
        for i, row in enumerate(rewards_data):

            # 处理图片路径
            image_path = row[5] or ''
            if image_path and not image_path.startswith('/uploads/'):
                # 如果图片路径不是以/uploads/开头，添加前缀
                if not image_path.startswith('http'):
                    image_path = f'/uploads/{image_path}'

            reward_dict = {
                'id': row[0],
                'name': row[1],
                'description': row[2] or '',
                'points_required': row[3],
                'stock': row[4],
                'image_path': image_path,
                'category': row[6],
                'available': True  # 移除库存检查，因为有些奖品可能没有库存字段
            }
            if manage_mode:
                reward_dict['is_active'] = row[7] if len(row) > 7 else 1
            rewards.append(reward_dict)
        # 获取所有分类（暂时返回默认分类）
        categories = ['默认']

        # 根据manage_mode返回不同格式的数据
        if manage_mode:
            result = {
                'success': True,
                'data': rewards  # 管理模式直接返回数组
            }
        else:
            result = {
                'success': True,
                'data': {
                    'rewards': rewards,
                    'categories': categories
                }
            }

        return jsonify(result)

    except Exception as e:
        error_msg = f'获取奖品列表失败: {str(e)}'
        _safe_print(f"[ERROR] {error_msg}")
        _safe_print(f"[ERROR] 异常详情: {type(e).__name__}: {e}")
        import traceback
        _safe_print(f"[ERROR] 堆栈跟踪: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'message': error_msg
        }), 500


@app.route('/api/exchange_reward', methods=['POST'])
def exchange_reward():
    """兑换奖品API"""
    try:
        data = request.get_json() or {}
        user_name = _get_current_feishu_user_name()
        reward_id = data.get('reward_id')

        if not user_name:
            return jsonify({
                'success': False,
                'message': '请先登录飞书后再兑换'
            }), 401

        if not reward_id:
            return jsonify({
                'success': False,
                'message': '参数不完整'
            }), 400
        user_esc = user_name.replace("'", "''")

        try:
            reward_sql = f"""
                SELECT 奖品, 奖品对应积分, 库存
                FROM 创新_奖品表
                WHERE id = {reward_id}
            """

            reward_info = sf_db(reward_sql)

            if not reward_info:
                raise Exception('奖品不存在或已下架')

            reward_name, points_required, stock = reward_info[0]

            if stock <= 0:
                raise Exception('奖品库存不足')

            current_year = datetime.now().year
            annual_points_sql = f"""
                SELECT COALESCE(SUM(最高分), 0) as annual_points
                FROM v_quanyuanchuangxin 
                WHERE 发起人 = '{user_esc}' 
                AND YEAR(发起时间) = {current_year}
                AND 最高分 IS NOT NULL AND 最高分 > 0
            """
            total_points_sql = f"""
                SELECT COALESCE(SUM(最高分), 0) as total_points
                FROM v_quanyuanchuangxin 
                WHERE 发起人 = '{user_esc}'
                AND 最高分 IS NOT NULL AND 最高分 > 0
            """
            spent_sql = f"""
                SELECT COALESCE(SUM(花费积分), 0) as spent_points
                FROM 创新_消费 
                WHERE 使用人 = '{user_esc}'
            """

            def extract_value(result):
                if not result:
                    return 0
                if isinstance(result, list):
                    if len(result) > 0:
                        if isinstance(result[0], (list, tuple)):
                            return result[0][0] if len(result[0]) > 0 else 0
                        else:
                            return result[0]
                    return 0
                if isinstance(result, (int, float)):
                    return result
                return 0

            annual_result = sf_db(annual_points_sql)
            total_result = sf_db(total_points_sql)
            spent_result = sf_db(spent_sql)

            annual_points = extract_value(annual_result)
            total_earned_points = extract_value(total_result)
            spent_points = extract_value(spent_result)

            current_points = total_earned_points - spent_points

            if current_points < points_required:
                raise Exception(f'积分不足，需要{points_required}积分，当前只有{current_points}积分')

            # 更新库存
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            update_stock_sql = f"""
                UPDATE 创新_奖品表
                SET 库存 = 库存 - 1
                WHERE id = {reward_id}
            """

            dui_db(update_stock_sql)

            # 添加消费记录到创新_消费表
            insert_consumption_sql = f"""
                INSERT INTO 创新_消费 (使用人, 日期, 花费积分, 备注)
                VALUES ('{user_esc}', '{current_time}', {points_required}, '兑换奖品「{reward_name}」')
            """

            dui_db(insert_consumption_sql)

            # 计算剩余积分
            new_points = current_points - points_required

            # 发送飞书通知
            try:
                message_service = get_message_service()
                message_service.send_exchange_notification(
                    user_name=user_name,
                    reward_name=reward_name,
                    points_spent=points_required,
                    remaining_points=new_points
                )
                message_service.send_exchange_notification_renli(
                    notice_name='蔡晶',
                    user_name=user_name,
                    reward_name=reward_name,
                    points_spent=points_required,
                    remaining_points=new_points
                )
                message_service.send_exchange_notification_renli(
                    notice_name='韩雅俊',
                    user_name=user_name,
                    reward_name=reward_name,
                    points_spent=points_required,
                    remaining_points=new_points
                )

            except Exception as notify_error:
                _safe_print(f"[WARNING] 飞书通知发送失败: {notify_error}")
                # 通知失败不影响兑换成功

            return jsonify({
                'success': True,
                'message': f'成功兑换「{reward_name}」！消耗{points_required}积分，剩余{new_points}积分。人力部门将收到通知并安排发放。'
            })

        except Exception as e:
            _safe_print(f"[ERROR] 兑换过程中发生错误: {e}")
            raise e


    except Exception as e:
        error_msg = str(e)
        _safe_print(f"[ERROR] 兑换失败: {error_msg}")
        return jsonify({
            'success': False,
            'message': error_msg
        }), 500


@app.route('/api/get_point_records', methods=['GET'])
def get_point_records():
    """获取积分记录API"""
    try:
        user_name = (request.args.get('user_name') or '').strip()
        if not user_name:
            user_name = _get_current_feishu_user_name()
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))

        if not user_name:
            return jsonify({
                'success': False,
                'message': '请先登录飞书后再查看积分记录'
            }), 401
        user_esc = user_name.replace("'", "''")

        # 构建积分记录：合并积分获得和消费记录
        records = []

        earn_sql = f"""
            SELECT
                v.编号 as 项目编号,
                CASE
                    WHEN ISNUMERIC(ISNULL(v.最高分, 0)) = 1 THEN CAST(ISNULL(v.最高分, 0) AS INT)
                    ELSE 0
                END as points,
                '获得' as type,
                '创新项目' as source,
                CONCAT('创新项目评分获得积分 - 项目编号:', v.编号, ' - ', COALESCE(v.标题, '未知项目')) as description,
                COALESCE(v.标题, '未知项目') as project_title
            FROM v_quanyuanchuangxin v
            WHERE v.发起人 = '{user_esc}'
              AND ISNUMERIC(ISNULL(v.最高分, 0)) = 1
              AND CAST(ISNULL(v.最高分, 0) AS INT) > 0
        """

        # 同时修改spend_sql以保持字段一致
        spend_sql = f"""
            SELECT 
             0 as 项目编号,
                -花费积分 as points,
                '消费' as type,
                '积分兑换' as source,
                COALESCE(备注, '积分消费') as description,


                '' as project_title
            FROM 创新_消费 
            WHERE 使用人 = '{user_esc}'
            AND 花费积分 > 0
            AND (备注 IS NULL OR 备注 != '积分汇总')
        """

        # 3. 合并查询并按时间排序
        combined_sql = f"""
            SELECT * FROM (
                {earn_sql}
                UNION ALL
                {spend_sql}
            ) AS combined_records
            ORDER BY 项目编号 DESC
        """

        # 执行查询
        records_data = sf_db(combined_sql)

        if records_data:
            # 分页处理
            total_records = len(records_data)
            start_index = (page - 1) * per_page
            end_index = start_index + per_page
            paginated_data = records_data[start_index:end_index]

            records = []
            for row in paginated_data:
                # 处理积分显示
                points = row[1] if row[1] is not None else 0
                record_type = row[2]
                source = row[3]
                description = row[4]

                project_id = row[0]
                project_title = row[5]
                record_time = sf_db(f"select 发起时间 from chuangxin_tibao1 where 编号='{project_id}'")

                # 格式化描述信息
                if source == '创新项目' and points > 0:
                    title_part = f" - {project_title}" if project_title else ""
                    if points >= 100:
                        description = f"创新项目A+级评分奖励 - 项目编号:{project_id} ({project_title}) (+{points}分)"
                    elif points >= 50:
                        description = f"创新项目A级评分奖励 - 项目编号:{project_id} ({project_title}) (+{points}分)"
                    elif points >= 30:
                        description = f"创新项目B级评分奖励 - 项目编号:{project_id} ({project_title}) (+{points}分)"
                    elif points >= 20:
                        description = f"创新项目C级评分奖励 - 项目编号:{project_id} ({project_title}) (+{points}分)"
                    elif points >= 10:
                        description = f"创新项目D级评分奖励 - 项目编号:{project_id} ({project_title}) (+{points}分)"
                    else:
                        description = f"创新项目评分奖励 - 项目编号:{project_id} ({project_title}) (+{points}分)"
                elif source == '积分兑换' and points < 0:
                    description = f"积分兑换消费 ({points}分) - 兑换奖品 {description.replace('积分消费', '').replace('积分兑换消费', '')}"

                records.append({
                    'points': points,
                    'type': record_type,
                    'source': source,
                    'description': description,
                    'record_time': record_time,
                    'formatted_points': f"+{points}" if points > 0 else str(points),
                    'type_class': 'text-success' if points > 0 else 'text-danger',
                    'icon': '📈' if points > 0 else '📉'
                })
        else:
            total_records = 0
            records = []

        # 统计口径与当前积分保持一致：获得积分来自v_quanyuanchuangxin，消费来自创新_消费
        stats = {'total_earned': 0, 'total_spent': 0, 'earn_count': 0, 'spend_count': 0}
        stats_sql = f"""
            SELECT
                COALESCE(SUM(CASE WHEN ISNUMERIC(ISNULL(最高分, 0)) = 1 THEN CAST(ISNULL(最高分, 0) AS INT) ELSE 0 END), 0) AS earned_sum,
                COALESCE(COUNT(CASE WHEN ISNUMERIC(ISNULL(最高分, 0)) = 1 AND CAST(ISNULL(最高分, 0) AS INT) > 0 THEN 1 END), 0) AS earned_count
            FROM v_quanyuanchuangxin
            WHERE 发起人 = '{user_esc}'
        """
        spent_stats_sql = f"""
            SELECT
                COALESCE(SUM(花费积分), 0) AS spent_sum,
                COALESCE(COUNT(1), 0) AS spent_count
            FROM 创新_消费
            WHERE 使用人 = '{user_esc}'
              AND 花费积分 > 0
              AND (备注 IS NULL OR 备注 != '积分汇总')
        """
        stats_rows = sf_db(stats_sql) or []
        spent_rows = sf_db(spent_stats_sql) or []
        if stats_rows and isinstance(stats_rows[0], (list, tuple)):
            stats['total_earned'] = int(stats_rows[0][0] or 0)
            stats['earn_count'] = int(stats_rows[0][1] or 0)
        if spent_rows and isinstance(spent_rows[0], (list, tuple)):
            stats['total_spent'] = int(spent_rows[0][0] or 0)
            stats['spend_count'] = int(spent_rows[0][1] or 0)

        return jsonify({
            'success': True,
            'data': {
                'records': records,
                'total': total_records,
                'page': page,
                'per_page': per_page,
                'total_pages': (total_records + per_page - 1) // per_page if total_records > 0 else 0,
                'stats': stats
            }
        })

    except Exception as e:
        _safe_print(f"[ERROR] 获取积分记录失败: {str(e)}")
        import traceback
        _safe_print(f"[ERROR] 堆栈跟踪: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'message': f'获取积分记录失败: {str(e)}'
        }), 500


@app.route('/api/export_innovations', methods=['GET'])
def export_innovations():
    """导出创新提案数据到Excel"""
    try:
        # 获取查询参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        operation_types = request.args.getlist('operation_types')

        _safe_print(f"📊 导出参数 - 开始日期: {start_date}, 结束日期: {end_date}, 操作类型: {operation_types}")

        # 部门映射
        department_mapping = {
            '孙洁': '运营一部',
            '刘蓉蓉': '运营三部',
            '侯梁': '运营六部',
            '宋亚倩 ': '运营二部',
            '王钰媛': '运营二部',
            '数据': 'AI部',
            'TK': 'TK项目',
            '技术': '技术部',
            '采购': '采购部',
            '美工': '视觉设计部',
            '人力': '人力行政部',
            '摄影': '摄影部',
            '财务': '财务部',
            '未知部门': '仓库'
            # 可以根据实际情况添加更多映射
        }

        # 构建日期过滤条件
        date_filter = ""
        if start_date:
            date_filter += f" AND v.发起时间 >= '{start_date}'"
        if end_date:
            date_filter += f" AND v.发起时间 <= '{end_date} 23:59:59'"

        # 构建操作类型过滤条件
        operation_filter = ""
        if operation_types:
            if '' in operation_types:
                operation_types_escaped = [f"'{op.replace("'", "''")}" for op in operation_types if op]
                if operation_types_escaped:
                    operation_filter = f" AND (t2_操作.操作类型 IN ({','.join(operation_types_escaped)}) OR t2_操作.操作类型 IS NULL OR t2_操作.操作类型 = '')"
                else:
                    operation_filter = " AND (t2_操作.操作类型 IS NULL OR t2_操作.操作类型 = '')"
            else:
                operation_types_escaped = [f"'{op.replace("'", "''")}" for op in operation_types]
                operation_filter = f" AND t2_操作.操作类型 IN ({','.join(operation_types_escaped)})"

        # 使用v_QuanYuanChuangXin视图获取项目得分数据
        export_sql = f"""
	WITH CTE AS (
    SELECT 
        ROW_NUMBER() OVER (PARTITION BY v.内容 ORDER BY v.最高分 DESC, v.编号) AS rn,  -- 按内容分组，分数高的优先
        ROW_NUMBER() OVER (ORDER BY v.编号) as 序号,
        v.发起人 as 提案人,
        v.部门 as 所在部门,
        ISNULL(v.创新类别, '') as 创新类别,
        v.标题 as 提案项目,
        ISNULL(v.内容, '') as 提案内容,
        ISNULL(v.解决方案, '') as 解决方案,
        ISNULL(v.图片详情, '') as 相关图片,
        ISNULL(CONVERT(varchar, v.截止时间, 23), '') as 期望完成时间,
        ISNULL(CONVERT(varchar, v.发起时间, 23), '') as 发起日期,
        ISNULL(STUFF((SELECT ', ' + 承接人
                       FROM chuangxin_liuzhuan1 t2_sub
                      WHERE t2_sub.项目编号 = v.编号
                      FOR XML PATH('')), 1, 2, ''),'') as 承接部门,
        ISNULL(STUFF((SELECT ', ' + 操作类型
                       FROM chuangxin_liuzhuan1 t2_操作
                      WHERE t2_操作.项目编号 = v.编号 AND 操作类型 IS NOT NULL AND 操作类型 != ''
                      FOR XML PATH('')), 1, 2, ''),'暂无操作') as 操作类型,
        ISNULL(CAST(v.最高分 AS varchar), '未评分') as 提案评分,
        ISNULL(STUFF((SELECT ', ' + 处理备注
                       FROM chuangxin_liuzhuan1 t2_备注
                      WHERE t2_备注.项目编号 = v.编号 AND 处理备注 IS NOT NULL AND 处理备注 != ''
                      FOR XML PATH('')), 1, 2, ''),'暂无备注') as 处理备注
    FROM v_QuanYuanChuangXin v
    WHERE 1=1 {date_filter}
)
SELECT 
    序号, 提案人, 所在部门, 创新类别, 提案项目, 提案内容, 
    解决方案, 相关图片, 期望完成时间, 发起日期, 
    承接部门, 操作类型, 提案评分, 处理备注
FROM CTE
WHERE rn = 1   -- 每个“提案内容”只保留最高分那条
ORDER BY 序号;
        """

        _safe_print(f"🔍 执行导出SQL查询...")
        export_data = sf_db(export_sql)

        if not export_data:
            return jsonify({
                'success': False,
                'message': '没有数据可导出'
            }), 404

        _safe_print(f"📊 查询到 {len(export_data)} 条记录")

        # 创建DataFrame
        columns = [
            '序号', '提案人', '所在部门', '创新类别', '提案项目',
            '提案内容', '解决方案', '相关图片', '期望完成时间', '发起日期',
            '承接部门', '操作类型', '提案评分', '处理备注'
        ]

        df = pd.DataFrame(export_data, columns=columns)

        def map_department(dept_name):
            if not dept_name or dept_name == '未知部门':
                return '仓储部'
            for key, value in department_mapping.items():
                if key in dept_name:
                    return value
            return dept_name

        df['所在部门'] = df['所在部门'].apply(map_department)

        df['期望完成时间'] = df['期望完成时间'].apply(
            lambda x: x if x and x.strip() else '未设置'
        )
        df['发起日期'] = df['发起日期'].apply(
            lambda x: x if x and x.strip() else '未知'
        )

        # 生成Excel文件
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        date_suffix = f"_{start_date}_to_{end_date}" if start_date and end_date else ""
        filename = f'创新提案管理_{current_time}{date_suffix}.xlsx'
        filepath = os.path.join(APP_CONFIG['export_folder'], filename)

        # 确保导出目录存在
        os.makedirs(APP_CONFIG['export_folder'], exist_ok=True)

        # 使用openpyxl创建Excel文件并嵌入图片
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from PIL import Image as PILImage
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = '创新提案数据'

        # 写入表头
        headers = ['序号', '提案人', '所在部门', '创新类别', '提案项目',
                   '提案内容', '解决方案', '相关图片', '期望完成时间', '发起日期',
                   '承接部门', '操作类型', '提案评分', '处理备注']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置列宽
        column_widths = {
            'A': 8,  # 序号
            'B': 12,  # 提案人
            'C': 15,  # 所在部门
            'D': 15,  # 创新类别
            'E': 25,  # 提案项目
            'F': 40,  # 提案内容
            'G': 30,  # 解决方案
            'H': 35,  # 相关图片
            'I': 20,  # 期望完成时间
            'J': 20,  # 发起日期
            'K': 20,  # 承接部门
            'L': 15,  # 操作类型
            'M': 12,  # 提案评分
            'N': 30  # 处理备注
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 写入数据并处理图片
        temp_files = []  # 记录临时文件，用于最后清理

        for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
            ws.row_dimensions[row_idx].height = 150  # 设置行高

            for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
                if col_name == '相关图片' and value and value.strip():
                    # 处理图片
                    try:
                        image_path_str = str(value).strip()

                        if image_path_str.lower().startswith('0x'):
                            try:
                                image_path_str = bytes.fromhex(image_path_str[2:]).decode('utf-8')
                            except (ValueError, UnicodeDecodeError):
                                pass

                        # 处理字节字符串前缀
                        if image_path_str.startswith("b'") and image_path_str.endswith("'"):
                            image_path_str = image_path_str[2:-1]

                        # 标准化路径
                        image_path_str = image_path_str.replace('\\\\', '\\').replace('//', '/')
                        image_path_str = image_path_str.strip('"\'')

                        # 尝试多个可能的路径
                        possible_paths = [
                            image_path_str,  # 原始路径
                            os.path.join(app.config['UPLOAD_FOLDER'], os.path.basename(image_path_str)),
                            os.path.join(APP_CONFIG['legacy_upload_folder'], os.path.basename(image_path_str)),
                            os.path.join('static', 'uploads', os.path.basename(image_path_str)),  # static/uploads/文件名
                            os.path.join('uploads', os.path.basename(image_path_str)),  # uploads/文件名
                            os.path.join(os.getcwd(), 'static', 'uploads', os.path.basename(image_path_str)),  # 绝对路径
                            os.path.join(os.getcwd(), 'uploads', os.path.basename(image_path_str))  # 绝对路径uploads
                        ]

                        image_found = False
                        for path in possible_paths:
                            if os.path.exists(path):

                                # 使用PIL处理图片
                                pil_image = PILImage.open(path)

                                # 转换图片模式
                                if pil_image.mode in ('RGBA', 'LA', 'P'):
                                    pil_image = pil_image.convert('RGB')

                                # 调整图片大小
                                pil_image = pil_image.resize((200, 150), PILImage.Resampling.LANCZOS)

                                # 保存临时图片
                                temp_path = f"temp_image_{row_idx}_{col_idx}.jpg"
                                pil_image.save(temp_path, 'JPEG', quality=85)
                                temp_files.append(temp_path)  # 记录临时文件

                                # 插入到Excel
                                img = OpenpyxlImage(temp_path)
                                img.width = 200
                                img.height = 150

                                cell_ref = f'{chr(64 + col_idx)}{row_idx}'
                                ws.add_image(img, cell_ref)

                                image_found = True
                                break

                        if not image_found:
                            ws.cell(row=row_idx, column=col_idx,
                                    value=f"图片文件不存在: {os.path.basename(image_path_str)}")

                    except Exception as img_error:
                        _safe_print(f"❌ 处理图片失败 {value}: {img_error}")
                        ws.cell(row=row_idx, column=col_idx, value=f"图片处理失败: {os.path.basename(str(value))}")
                else:
                    # 写入普通数据
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # 保存Excel文件
        wb.save(filepath)

        # 清理临时文件
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)

            except Exception as cleanup_error:
                _safe_print(f"⚠️ 删除临时文件失败 {temp_file}: {cleanup_error}")

        _safe_print(f"✅ Excel文件生成成功，文件名: {filename}")

        # 返回文件
        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        _safe_print(f"❌ 导出失败: {e}")
        import traceback
        _safe_print(f"📋 错误堆栈: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'message': f'导出失败: {str(e)}'
        }), 500


@app.route('/api/get_statistics', methods=['GET'])
def get_statistics():
    """获取统计数据API"""
    try:
        # 获取各状态统计
        total_count = sf_db("SELECT COUNT(*) FROM chuangxin_tibao1", single=True)
        pending_count = sf_db("SELECT COUNT(*) FROM chuangxin_tibao1 WHERE 状态 = '待承接'", single=True)
        processing_count = sf_db("SELECT COUNT(*) FROM chuangxin_tibao1 WHERE 状态 = '进行中'", single=True)
        completed_count = sf_db("SELECT COUNT(*) FROM chuangxin_tibao1 WHERE 状态 = '已完成'", single=True)

        # 获取平均评分
        avg_score_result = sf_db("SELECT AVG(CAST(得分 AS FLOAT)) FROM chuangxin_tibao1 WHERE 得分 > 0", single=True)
        avg_score = round(avg_score_result, 1) if avg_score_result else 0

        # 获取部门统计（从流转表获取）
        department_stats_data = sf_db("""
                                      SELECT f.承接人, COUNT(DISTINCT f.项目编号) as count
                                      FROM chuangxin_liuzhuan1 f
                                      GROUP BY f.承接人
                                      ORDER BY count DESC
                                      """)
        department_stats = [{'department': row[0], 'count': row[1]} for row in department_stats_data]

        # 获取评分分布
        score_distribution_data = sf_db("""
                                        SELECT CASE
                                                   WHEN 得分 >= 9 THEN '优秀(9-10分)'
                                                   WHEN 得分 >= 7 THEN '良好(7-8分)'
                                                   WHEN 得分 >= 4 THEN '一般(4-6分)'
                                                   WHEN 得分 > 0 THEN '较差(1-3分)'
                                                   ELSE '未评分'
                                                   END as score_range,
                                               COUNT(*) as count
                                        FROM chuangxin_tibao1
                                        GROUP BY
                                            CASE
                                            WHEN 得分 >= 9 THEN '优秀(9-10分)'
                                            WHEN 得分 >= 7 THEN '良好(7-8分)'
                                            WHEN 得分 >= 4 THEN '一般(4-6分)'
                                            WHEN 得分 > 0 THEN '较差(1-3分)'
                                            ELSE '未评分'
                                        END
                                        ORDER BY count DESC
                                        """)
        score_distribution = [{'range': row[0], 'count': row[1]} for row in score_distribution_data]

        return jsonify({
            'success': True,
            'data': {
                'total_count': total_count,
                'pending_count': pending_count,
                'processing_count': processing_count,
                'completed_count': completed_count,
                'avg_score': avg_score,
                'department_stats': department_stats,
                'score_distribution': score_distribution
            }
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'获取统计数据失败: {str(e)}'
        }), 500


@app.route('/api/export_data', methods=['GET'])
def export_data():
    """导出数据API"""
    try:
        sql = """
              SELECT 编号 as '项目ID', 标题 as '项目标题', 内容 as '项目内容', 发起人 as '发起人', 发起时间 as '创建时间', 截止时间 as '截止时间', 得分 as '评分', 状态 as '状态', 处理人 as '处理人', 处理时间 as '处理时间', 处理备注 as '处理备注'
              FROM chuangxin_tibao1
              ORDER BY 发起时间 DESC
              """

        data = sf_db(sql)

        # 转换为DataFrame
        columns = ['项目ID', '项目标题', '项目内容', '发起人', '创建时间', '截止时间', '评分', '状态',
                   '处理人', '处理时间', '处理备注']
        df = pd.DataFrame(data, columns=columns)

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'创新项目数据_{timestamp}.xlsx'
        filepath = os.path.join(APP_CONFIG['export_folder'], filename)

        # 导出Excel
        df.to_excel(filepath, index=False, engine='openpyxl')

        return jsonify({
            'success': True,
            'message': '数据导出成功',
            'filename': filename
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'导出失败: {str(e)}'
        }), 500


# 奖品管理API（需要管理员权限）
@app.route('/api/add_reward', methods=['POST'])
def add_reward():
    """添加新奖品API - 所有用户可访问"""
    try:
        data = request.get_json()
        name = data.get('name')
        description = data.get('description', '')
        points_required = data.get('points_required')
        stock = data.get('stock', -1)
        category = data.get('category', '其他')

        if not name or not points_required:
            return jsonify({'success': False, 'message': '奖品名称和积分要求不能为空'}), 400

        # 使用中文字段名的SQL语句
        sql = f"""
            INSERT INTO 创新_奖品表 (奖品, 商品描述, 奖品对应积分, 库存, 商品分类)
            VALUES ('{name}', '{description}', {points_required}, {stock}, '{category}')
        """

        dui_db(sql)

        return jsonify({
            'success': True,
            'message': '奖品添加成功'
        })

    except Exception as e:

        return jsonify({
            'success': False,
            'message': f'添加奖品失败: {str(e)}'
        }), 500


@app.route('/api/upload_reward_image', methods=['POST'])
def upload_reward_image():
    """上传奖品图片API - 所有用户可访问"""
    try:
        if 'image' not in request.files:
            return jsonify({'success': False, 'message': '没有选择文件'}), 400

        file = request.files['image']
        reward_id = request.form.get('reward_id')

        if not reward_id:
            return jsonify({'success': False, 'message': '缺少奖品ID'}), 400

        if file.filename == '':
            return jsonify({'success': False, 'message': '没有选择文件'}), 400

        if file:
            # 生成安全的文件名
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{filename}"

            # 保存文件
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'rewards', filename)
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            file.save(filepath)

            # 更新数据库中的图片路径（修正字段名）
            relative_path = f"rewards/{filename}"
            sql = f"UPDATE 创新_奖品表 SET 奖励图片 = '{relative_path}' WHERE id = {reward_id}"
            dui_db(sql)

            return jsonify({
                'success': True,
                'message': '图片上传成功',
                'image_path': relative_path
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'上传失败: {str(e)}'
        }), 500


@app.route('/api/update_reward', methods=['POST'])
def update_reward():
    """更新奖品API - 所有用户可访问"""
    try:
        data = request.get_json()
        reward_id = data.get('reward_id')
        name = data.get('name')
        description = data.get('description', '')
        points_required = data.get('points_required')
        stock = data.get('stock', -1)
        category = data.get('category', '其他')
        is_active = data.get('is_active', True)

        if not reward_id or not name or not points_required:
            return jsonify({'success': False, 'message': '缺少必要参数'}), 400

        # 使用中文字段名的SQL语句
        sql = f"""
            UPDATE 创新_奖品表 
            SET 奖品 = '{name}', 商品描述 = '{description}', 奖品对应积分 = {points_required}, 
                库存 = {stock}, 商品分类 = '{category}'
            WHERE id = {reward_id}
        """

        dui_db(sql)

        return jsonify({
            'success': True,
            'message': '奖品更新成功'
        })

    except Exception as e:

        return jsonify({
            'success': False,
            'message': f'更新奖品失败: {str(e)}'
        }), 500


@app.route('/api/delete_reward', methods=['POST'])
def delete_reward():
    """删除奖品API - 所有用户可访问"""
    try:
        data = request.get_json()
        reward_id = data.get('reward_id')

        if not reward_id:
            return jsonify({'success': False, 'message': '缺少奖品ID'}), 400

        # 检查是否有兑换记录（修正表名和字段名）
        check_sql = f"SELECT COUNT(*) FROM 兑换记录表 WHERE reward_id = {reward_id}"
        try:
            exchange_count = sf_db(check_sql, single=True)
        except:
            # 如果兑换记录表不存在或字段名不对，设为0
            exchange_count = 0

        # 直接删除奖品（因为创新_奖品表可能没有is_active字段）
        sql = f"DELETE FROM 创新_奖品表 WHERE id = {reward_id}"
        message = '奖品删除成功'

        # 如果有兑换记录，提示但仍然删除
        if exchange_count > 0:
            message = '奖品删除成功（注意：该奖品存在兑换记录）'

        dui_db(sql)

        return jsonify({
            'success': True,
            'message': message
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'删除奖品失败: {str(e)}'
        }), 500


@app.route('/api/add_reward_with_image', methods=['POST'])
def add_reward_with_image():
    """添加带图片的奖品API - 所有用户可访问"""
    try:
        # 获取表单数据
        name = request.form.get('name')
        description = request.form.get('description', '')
        points_required = request.form.get('points_required')
        stock = request.form.get('stock', -1)
        category = request.form.get('category', '其他')

        if not name or not points_required:
            return jsonify({'success': False, 'message': '奖品名称和积分要求不能为空'}), 400

        # 处理图片上传
        image_path = None
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename:
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"

                # 确保rewards目录存在
                rewards_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'rewards')
                os.makedirs(rewards_dir, exist_ok=True)

                filepath = os.path.join(rewards_dir, filename)
                file.save(filepath)
                image_path = f"rewards/{filename}"

        # 插入奖品信息 - 使用正确的中文字段名
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sql = f"""
            INSERT INTO 创新_奖品表 (奖品, 奖品对应积分, 库存, 奖励图片)
            VALUES ('{name}', {points_required}, {stock}, '{image_path or ''}')
        """

        dui_db(sql)

        return jsonify({
            'success': True,
            'message': '奖品添加成功',
            'image_path': image_path
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'添加奖品失败: {str(e)}'
        }), 500


def add_points_consumption(user, points, note):
    """添加积分消费记录"""
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sql = f"""
        INSERT INTO 创新_消费 (使用人, 日期, 花费积分, 备注)
        VALUES ('{user}', '{current_time}', {points}, '{note}')
    """
    dui_db(sql)


def add_reward_item(reward_name, points_required, stock):
    """添加奖品项目"""
    sql = f"""
        INSERT INTO 创新_奖品表 (奖品, 奖品对应积分, 库存)
        VALUES ('{reward_name}', {points_required}, {stock})
    """
    dui_db(sql)


@app.route('/innovation_star_media/<path:filename>')
def innovation_star_media_file(filename):
    """提供创新星主场图片和视频访问。"""
    if not session.get('feishu_user_id'):
        return jsonify({'error': '请先登录'}), 401
    safe_name = os.path.basename(str(filename or '').replace('/', os.sep))
    if not safe_name or safe_name != filename:
        return jsonify({'error': '文件不存在'}), 404
    try:
        return send_from_directory(
            INNOVATION_STAR_MEDIA_FOLDER,
            safe_name,
            as_attachment=False,
            conditional=True,
        )
    except Exception:
        return jsonify({'error': '文件不存在'}), 404


# 在文件末尾添加图片访问路由
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    """提供上传文件的访问"""
    try:
        # 支持子目录访问
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except Exception as e:
        _safe_print(f"❌ 访问文件失败: {e}")
        return jsonify({'error': '文件不存在'}), 404


@app.route('/like_images/<path:filename>')
def like_image_file(filename):
    try:
        return send_from_directory(LIKE_IMAGE_FOLDER, filename)
    except Exception:
        return jsonify({'error': '文件不存在'}), 404


# 添加专门的奖品图片路由
@app.route('/static/rewards/<filename>')
def reward_image(filename):
    """提供奖品图片的访问"""
    try:
        return send_from_directory('static/rewards', filename)
    except Exception as e:
        _safe_print(f"❌ 访问奖品图片失败: {e}")
        return jsonify({'error': '图片不存在'}), 404


@app.route('/api/test_simple', methods=['GET'])
def test_simple():
    """简单测试API"""
    return jsonify({'status': 'ok', 'message': 'API working'})


@app.route('/api/get_dashboard_data', methods=['GET'])
def get_dashboard_data():
    """获取数字看板数据API"""
    try:
        # 获取查询参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        month = request.args.get('month')  # 保留兼容性

        # 构建日期过滤条件 - 优先使用日期范围查询
        date_filter = ""
        view_date_filter = ""

        # 如果提供了日期范围，使用日期范围查询
        if start_date or end_date:
            if start_date:
                date_filter += f" AND t1.发起时间 >= '{start_date}'"
                view_date_filter += f" AND 发起时间 >= '{start_date}'"
            if end_date:
                date_filter += f" AND t1.发起时间 <= '{end_date} 23:59:59'"
                view_date_filter += f" AND 发起时间 <= '{end_date} 23:59:59'"
        # 如果没有日期范围但有月份参数，使用月份查询（向后兼容）
        elif month:
            date_filter += f" AND FORMAT(t1.发起时间, 'yyyy-MM') = '{month}'"
            view_date_filter += f" AND FORMAT(发起时间, 'yyyy-MM') = '{month}'"

        # 1. 基础统计
        total_sql = f"SELECT COUNT(*) FROM chuangxin_tibao1 t1 WHERE 1=1 {date_filter}"
        total_count = sf_db(total_sql, single=True) or 0

        # 采纳提案量
        adopted_sql = f"""
            SELECT COUNT(DISTINCT t1.编号)
            FROM chuangxin_tibao1 t1
            JOIN chuangxin_liuzhuan1 t2 ON t1.编号 = t2.项目编号
            WHERE (
                CHARINDEX('立即执行', ISNULL(t2.操作类型, '')) > 0
                OR CHARINDEX('暂缓执行', ISNULL(t2.操作类型, '')) > 0
                OR CHARINDEX('已完成', ISNULL(t2.操作类型, '')) > 0
            ) {date_filter}
        """
        adopted_count = sf_db(adopted_sql, single=True) or 0

        # 提案有效率
        effectiveness_rate = round((adopted_count / total_count * 100), 2) if total_count > 0 else 0

        # 2. 部门提案积分排名
        dept_sql = f"""
            select top 20 部门 department, sum(最高分) total_score from v_QuanYuanChuangXin 
            where 1=1 {view_date_filter} and 部门 <>'Tk_离职'
            group by 部门 
            having sum(最高分)>0
            order by sum(最高分) desc
        """
        dept_raw = sf_db(dept_sql)
        top_departments = []
        if dept_raw and len(dept_raw) > 0:
            for row in dept_raw:
                top_departments.append({'name': row[0], 'score': round(row[1], 1)})
            top_dept = top_departments[0]
        else:
            top_dept = {'name': '暂无数据', 'score': 0}
            top_departments = [top_dept]

        # 3. 个人积分前三名
        person_sql = f"""
            select top 3 发起人 name , 部门 department, sum(最高分) total_score 
            from v_QuanYuanChuangXin
            where 1=1 {view_date_filter}
            group by 发起人, 部门
            order by sum(最高分) desc
        """
        person_raw = sf_db(person_sql)

        top_persons = []
        if person_raw and len(person_raw) > 0:
            for row in person_raw:
                person_data = {
                    'name': row[0],
                    'department': row[1],
                    'score': round(row[2], 1)
                }
                top_persons.append(person_data)
        else:
            top_persons = [{'name': '暂无数据', 'department': '未知部门', 'score': 0}]

        top_person = top_persons[0] if top_persons else {'name': '暂无', 'department': '', 'score': 0}

        # 4. 创新类别分布
        category_sql = f"""
            SELECT TOP 3
                ISNULL(创新类别, '未分类') as 创新类别,
                COUNT(*) as 数量
            FROM chuangxin_tibao1 t1
            WHERE 1=1 {date_filter}
            GROUP BY 创新类别
            ORDER BY 数量 DESC
        """
        category_raw = sf_db(category_sql)
        categories = []
        if category_raw and len(category_raw) > 0:
            for row in category_raw:
                categories.append({'name': row[0], 'count': row[1]})

        # 5. 月度趋势
        if month:
            monthly_sql = f"""
                SELECT
                    FORMAT(发起时间, 'yyyy-MM') as month,
                    COUNT(*) as count
                FROM chuangxin_tibao1 t1
                WHERE FORMAT(t1.发起时间, 'yyyy-MM') = '{month}'
                GROUP BY FORMAT(发起时间, 'yyyy-MM')
                ORDER BY month
            """
        else:
            monthly_sql = f"""
                SELECT
                    FORMAT(发起时间, 'yyyy-MM') as month,
                    COUNT(*) as count
                FROM chuangxin_tibao1 t1
                WHERE 发起时间 >= DATEADD(month, -6, GETDATE()) {date_filter}
                GROUP BY FORMAT(发起时间, 'yyyy-MM')
                ORDER BY month
            """
        monthly_raw = sf_db(monthly_sql)
        monthly_trend = []
        if monthly_raw and len(monthly_raw) > 0:
            for row in monthly_raw:
                monthly_trend.append({'month': row[0], 'count': row[1]})
        # 6.个人提案数量前三
        personal_stats_sql = f"""
                select top 3 姓名,部门,提案数量,提案有效量,提案有效率
                from(select 发起人 姓名,部门 部门,count(*) 提案数量, sum(case when 是否通过='Y'then 1 else 0 end ) 提案有效量,sum(case when 是否通过='Y'then 1 else 0 end )/(count(*)+0.0) 提案有效率 
                     from v_QuanYuanChuangXin v
                     WHERE 1=1 {date_filter.replace('t1.发起时间', 'v.发起时间')}
                     group by 发起人,部门) t
                order by 提案数量 desc
                               """
        personal_stats_raw = sf_db(personal_stats_sql)

        # 转换为字典格式
        top_proposal_counts = []
        if personal_stats_raw:
            for row in personal_stats_raw:
                top_proposal_counts.append({
                    'name': row[0],
                    'department': row[1],
                    'count': row[2],
                    'effective_count': row[3]
                })

        # 7.个人提案最高分
        highest_score_sql = f"""
                SELECT 发起人, 部门, 最高分 AS 分数, COUNT(*) AS 次数
                FROM v_QuanYuanChuangXin v
                WHERE 1=1
                  {date_filter.replace('t1.发起时间', 'v.发起时间')}
                GROUP BY 发起人, 部门, 最高分
                HAVING 最高分 = (
                    SELECT MAX(最高分) FROM v_QuanYuanChuangXin v2
                    WHERE 1=1 {date_filter.replace('t1.发起时间', 'v2.发起时间')}
                );
                               """
        highest_score_raw = sf_db(highest_score_sql)

        # 转换为字典格式
        top_highest_scores = []
        if highest_score_raw:
            for row in highest_score_raw:
                top_highest_scores.append({
                    'name': row[0],
                    'department': row[1],
                    'score': round(row[2], 1) if row[2] else 0,
                    'count': row[3] if len(row) > 3 else 1
                })

        response_data = {
            'success': True,
            'data': {
                'basic_stats': {
                    'total_count': total_count,
                    'adopted_count': adopted_count,
                    'effectiveness_rate': effectiveness_rate
                },
                'top_department': top_dept,
                'top_departments': top_departments,
                'top_person': top_person,
                'top_persons': top_persons,
                'categories': categories,
                'top_proposal_counts': top_proposal_counts,
                'top_highest_scores': top_highest_scores,
                'monthly_trend': monthly_trend
            }
        }

        return jsonify(response_data)

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'获取看板数据失败: {str(e)}'
        }), 500


@app.route('/api/get_innovation_compare', methods=['GET'])
def get_innovation_compare():
    """获取两条提案用于内容对比（标题/内容/解决方案/发起人/发起时间）"""
    try:
        current_bianhao = str(request.args.get('current_bianhao') or '').strip()
        target_bianhao = str(request.args.get('target_bianhao') or '').strip()
        if not current_bianhao or not target_bianhao:
            return jsonify({'success': False, 'message': '缺少对比编号参数'}), 400

        def _fetch_one(bianhao_text: str):
            b = bianhao_text.replace("'", "''")
            sql = f"""
                SELECT TOP 1
                    CAST(t.编号 AS NVARCHAR(100)) AS 编号,
                    ISNULL(CAST(t.标题 AS NVARCHAR(MAX)), '') AS 标题,
                    ISNULL(CAST(t.内容 AS NVARCHAR(MAX)), '') AS 内容,
                    ISNULL(CAST(t.解决方案 AS NVARCHAR(MAX)), '') AS 解决方案,
                    ISNULL(CAST(t.发起人 AS NVARCHAR(200)), '') AS 发起人,
                    t.发起时间
                FROM chuangxin_tibao1 t
                WHERE CAST(t.编号 AS NVARCHAR(100)) = '{b}'
                ORDER BY t.编号 DESC
            """
            rows = sf_db(sql) or []
            if not rows:
                return None
            r = rows[0]
            return {
                'bianhao': str(r[0]) if len(r) > 0 and r[0] is not None else bianhao_text,
                'title': str(r[1]) if len(r) > 1 and r[1] is not None else '',
                'content': str(r[2]) if len(r) > 2 and r[2] is not None else '',
                'solution': str(r[3]) if len(r) > 3 and r[3] is not None else '',
                'initiator': str(r[4]) if len(r) > 4 and r[4] is not None else '',
                'initiation_time': str(r[5]) if len(r) > 5 and r[5] is not None else '',
            }

        current_item = _fetch_one(current_bianhao)
        target_item = _fetch_one(target_bianhao)
        if not current_item or not target_item:
            return jsonify({'success': False, 'message': '未找到待对比提案'}), 404

        return jsonify({
            'success': True,
            'data': {
                'current': current_item,
                'target': target_item
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取对比数据失败: {str(e)}'}), 500


@app.route('/api/export_statistics', methods=['GET'])
def export_statistics():
    """导出统计数据到Excel"""
    try:
        # 获取查询参数
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        _safe_print(f"📊 统计导出参数 - 开始日期: {start_date}, 结束日期: {end_date}")

        # 构建日期过滤条件
        date_filter = ""
        if start_date:
            date_filter += f" AND t1.发起时间 >= '{start_date}'"
        if end_date:
            date_filter += f" AND t1.发起时间 <= '{end_date} 23:59:59'"

        # 1. 基础统计
        total_sql = f"SELECT COUNT(*) FROM chuangxin_tibao1 t1 WHERE 1=1 {date_filter}"
        total_count = sf_db(total_sql)[0] if sf_db(total_sql) else 0

        # 采纳提案量（含立即执行、暂缓执行、完成状态）
        adopted_sql = f"""
            SELECT COUNT(DISTINCT t1.编号) 
            FROM chuangxin_tibao1 t1 
            JOIN chuangxin_liuzhuan1 t2 ON t1.编号 = t2.项目编号 
            WHERE (
                CHARINDEX('立即执行', ISNULL(t2.操作类型, '')) > 0
                OR CHARINDEX('暂缓执行', ISNULL(t2.操作类型, '')) > 0
                OR CHARINDEX('已完成', ISNULL(t2.操作类型, '')) > 0
            ) {date_filter}
        """

        adopted_count = sf_db(adopted_sql)[0] if sf_db(adopted_sql) else 0

        # 提案有效率
        effectiveness_rate = round((adopted_count / total_count * 100), 2) if total_count > 0 else 0

        dept_stats_sql = f"""
                SELECT top 3 部门,总提报数量,有效提案数量,round(有效提案数量/(总提报数量+0.0),3) 有效比例
FROM (select 部门,count(*) 总提报数量,sum(case when 是否通过='Y' then 1 else 0 end) 有效提案数量
        from v_quanyuanchuangxin v
        WHERE 1=1 {date_filter.replace('t1.发起时间', 'v.发起时间')}
       group by 部门) t
 ORDER BY 总提报数量 desc

               """
        dept_stats_raw = sf_db(dept_stats_sql)

        # 转换为列表格式
        dept_stats = []
        if dept_stats_raw:
            for row in dept_stats_raw:
                dept_stats.append((row[0], row[1], row[2]))  # 部门, 总提报数量, 有效提案数量

        # 3. 部门积分统计
        dept_points_sql = f"""
                select top 1 部门,sum(最高分) 得分
from v_quanyuanchuangxin v
WHERE 1=1 {date_filter.replace('t1.发起时间', 'v.发起时间')}
group by 部门
order by 得分 desc
               """
        dept_points_raw = sf_db(dept_points_sql)

        # 转换为列表格式
        dept_points = []
        if dept_points_raw:
            for row in dept_points_raw:
                dept_points.append((row[0], row[1]))  # 部门, 得分

        # 4. 个人提案数量TOP3
        personal_stats_sql = f"""
select top 3 姓名,部门,提案数量,提案有效量,提案有效率
from(select 发起人 姓名,部门 部门,count(*) 提案数量, sum(case when 是否通过='Y'then 1 else 0 end ) 提案有效量,sum(case when 是否通过='Y'then 1 else 0 end )/(count(*)+0.0) 提案有效率 
     from v_QuanYuanChuangXin v
     WHERE 1=1 {date_filter.replace('t1.发起时间', 'v.发起时间')}
     group by 发起人,部门) t
order by 提案数量 desc
               """
        personal_stats_raw = sf_db(personal_stats_sql)

        # 转换为列表格式
        personal_stats = []
        if personal_stats_raw:
            for row in personal_stats_raw:
                personal_stats.append((row[0], row[1], row[2], row[3]))  # 姓名, 部门, 提案数量, 提案有效量

        # 5. 个人积分总分TOP3
        personal_points_sql = f"""
select top 3 姓名, 部门, 分数
from (select 发起人 姓名,部门,sum(最高分) 分数 
from v_QuanYuanChuangXin v
WHERE 1=1 {date_filter.replace('t1.发起时间', 'v.发起时间')}
group by 发起人,部门) t
order by 分数 desc
               """
        personal_points_raw = sf_db(personal_points_sql)

        # 转换为列表格式
        personal_points = []
        if personal_points_raw:
            for row in personal_points_raw:
                personal_points.append((row[0], row[1], row[2]))  # 姓名, 部门, 分数

        # 6. 提案单项得分最高
        highest_score_sql = f"""

select  top 3 发起人 姓名, 部门, 最高分 分数
from v_QuanYuanChuangXin v
WHERE 1=1 {date_filter.replace('t1.发起时间', 'v.发起时间')}
order by 最高分 desc
               """
        highest_scores_raw = sf_db(highest_score_sql)

        # 转换为列表格式
        highest_scores = []
        if highest_scores_raw:
            for row in highest_scores_raw:
                highest_scores.append({'发起人': row[0], '部门': row[1], '分数': row[2]})

        # 7. 创新类别分布
        category_sql = f"""
                   SELECT 
                       ISNULL(创新类别, '未分类') as 创新类别,
                       COUNT(*) as 数量
                   FROM chuangxin_tibao1 t1
                   WHERE 1=1 {date_filter}
                   GROUP BY 创新类别
                   ORDER BY 数量 DESC
               """
        category_stats_raw = sf_db(category_sql)

        # 转换为列表格式
        category_stats = []
        if category_stats_raw:
            for row in category_stats_raw:
                category_stats.append((row[0], row[1]))  # 创新类别, 数量

        wb = Workbook()
        ws = wb.active
        ws.title = '创新提案统计报告'

        # 设置样式
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        sub_header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        row = 1

        # 1. 基础统计
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = '1.基础统计'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_alignment
        row += 1

        # 表头
        headers = ['指标', '数值', '说明', '']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = sub_header_fill
            cell.alignment = center_alignment
            cell.border = border
        row += 1

        # 数据行
        basic_stats = [
            ['总提案量', f'{total_count}条', f'统计周期：{start_date or "开始"}-{end_date or "结束"}'],
            ['采纳提案量', f'{adopted_count}条', '含立即执行、暂缓执行、完成状态'],
            ['提案有效率', f'{effectiveness_rate}%', f'采纳量/总量：{adopted_count}/{total_count}']
        ]

        for stat in basic_stats:
            for col, value in enumerate(stat, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = center_alignment
            row += 1

        row += 2

        # 2. 部门提案数量排名
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = '2.部门提案数量排名'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_alignment
        row += 1

        dept_headers = ['排名', '部门', '数量', '提案有效量', '提案有效率']
        for col, header in enumerate(dept_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = sub_header_fill
            cell.alignment = center_alignment
            cell.border = border
        row += 1

        for i, dept in enumerate(dept_stats[:3], 1):
            dept_name, count, effective_count = dept
            effective_rate = round((effective_count / count * 100), 2) if count > 0 else 0
            rank_names = ['第一名', '第二名', '第三名']

            data = [rank_names[i - 1], dept_name, f'{count}条', f'{effective_count}条', f'{effective_rate}%']
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = center_alignment
            row += 1

        row += 2

        # 3. 部门提案积分总分第一名
        if dept_points:
            top_dept = dept_points[0]
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = f'3.部门提案积分总分第一名    {top_dept[0]}    {top_dept[1]}分'
            ws[f'A{row}'].font = title_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = center_alignment
            row += 2

        # 4. 个人提案数量TOP3
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = '4.个人提案数量TOP3'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_alignment
        row += 1

        personal_headers = ['姓名', '部门', '提案数量', '提案有效量', '提案有效率']
        for col, header in enumerate(personal_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = sub_header_fill
            cell.alignment = center_alignment
            cell.border = border
        row += 1

        for person in personal_stats:
            name, dept, count, effective_count = person
            effective_rate = round((effective_count / count * 100), 2) if count > 0 else 0

            data = [name, dept, f'{count}条', f'{effective_count}条', f'{effective_rate}%']
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = center_alignment
            row += 1

        row += 2

        # 5. 个人积分总分TOP3
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = '5.个人积分总分TOP3'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_alignment
        row += 1

        points_headers = ['姓名', '部门', '分数', '']
        for col, header in enumerate(points_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = sub_header_fill
            cell.alignment = center_alignment
            cell.border = border
        row += 1

        for person in personal_points:
            name, dept, points = person

            data = [name, dept, f'{points}分', '']
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = center_alignment
            row += 1

        row += 2

        # 6. 提案单项得分最高
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = '6.提案单项得分最高'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_alignment
        row += 1

        score_headers = ['姓名', '部门', '分数', '']
        for col, header in enumerate(score_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = sub_header_fill
            cell.alignment = center_alignment
            cell.border = border
        row += 1

        # 获取最高分
        if highest_scores:
            max_score = highest_scores[0]['分数']
            top_scorers = [score for score in highest_scores if score['分数'] == max_score]

            for scorer in top_scorers:
                name = scorer['发起人']
                dept = scorer['部门']
                score = scorer['分数']
                data = [name, dept, str(score), '']
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    cell.alignment = center_alignment
                row += 1

        row += 2

        # 7. 创新类别分布
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = '7.创新类别分布'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_alignment
        row += 1

        for category in category_stats:
            category_name, count = category
            percentage = round((count / total_count * 100), 2) if total_count > 0 else 0

            data = [category_name, f'{count}条', f'{percentage}%', '']
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = center_alignment
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15

        # 生成文件
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        date_suffix = f"_{start_date}_to_{end_date}" if start_date and end_date else ""
        filename = f'创新提案统计报告_{current_time}{date_suffix}.xlsx'
        filepath = os.path.join(APP_CONFIG['export_folder'], filename)

        # 确保导出目录存在
        os.makedirs(APP_CONFIG['export_folder'], exist_ok=True)

        wb.save(filepath)

        _safe_print(f"📊 统计报告导出成功: {filename}")

        # 返回文件
        return send_file(filepath, as_attachment=True, download_name=filename)

    except Exception as e:
        _safe_print(f"💥 统计数据导出失败: {e}")
        import traceback
        _safe_print(f"📋 错误堆栈: {traceback.format_exc()}")

        return jsonify({
            'success': False,
            'message': f'统计数据导出失败: {str(e)}'
        }), 500


if __name__ == '__main__':
    # _safe_print("🚀 正在启动创新管理系统...")
    # _safe_print("📊 数据库连接: SQL Server")
    # _safe_print("🔗 数据库操作: bjc.py (sf_db, dui_db)")

    try:
        message_service = MessageService('company1')
        if hasattr(message_service, 'test_connection') and message_service.test_connection():
            _safe_print("✅ 消息服务测试成功")
        else:
            _safe_print("⚠️  消息服务测试失败，请检查配置")
    except Exception as e:
        _safe_print(f"⚠️  消息服务测试异常: {e}")

    # _safe_print("🌐 服务器启动中...")
    # _safe_print("📱 访问地址: http://localhost:5000")

    app.run(host='0.0.0.0', port=5000, debug=True)
